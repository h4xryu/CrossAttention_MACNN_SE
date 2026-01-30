"""
Automated Experiment Script (Intra-Patient Version)

Intra-patient 패러다임(Stratified K-Fold)을 적용하여 실험을 수행합니다.
모든 환자의 데이터를 섞은 후 클래스 비율을 유지하며 5-Fold 교차 검증을 진행합니다.
학습 완료 후 자동으로 Evaluation을 수행하여 상세 엑셀 리포트를 생성합니다.

사용법:
    python main_autoexp_intra.py
"""

import os
import copy
import time
from datetime import datetime
from collections import defaultdict

import numpy as np
import torch
from torch.utils.data import DataLoader
from sklearn.model_selection import StratifiedKFold, train_test_split
from sklearn.metrics import confusion_matrix
from openpyxl import Workbook

import config
from src.registry import MODEL_REGISTRY, LOSS_REGISTRY, DATASET_REGISTRY
from src import models, losses, datasets
from src.optimizers import build_optimizer
from src.schedulers import build_scheduler
from src.trainer import Trainer
from src.utils import ExcelResultWriter, CumulativeExcelWriter
from utils import set_seed, load_or_extract_data


# =============================================================================
# 실험 설정
# =============================================================================

# Intra-patient 설정
N_FOLDS = 5  # 5-Fold CV
ALL_RECORDS = config.DS1_TRAIN + config.DS1_VALID + config.DS2_TEST # 전체 데이터 통합

# 실험할 파라미터 그리드
EXPERIMENT_GRID = {
    # Fusion type 실험
    "fusion_type": ["concat", "mhca", None],
    
    # MHCA용 num_heads
    "fusion_num_heads": [1],
}

# Fusion 실험용 설정
FUSION_EXP_CONFIG = {
    "lead": 1,
    "rr_dim": 7,
    "dataset_name": "ecg_standard",
    "rr_feature_option": "opt1",
}

EXP_EPOCHS = 50
OUTPUT_DIR = "./autoexp_results_intra_202602/"
CLASSES = config.CLASSES  # ["N", "S", "V", "F"]


# =============================================================================
# 유틸리티 함수
# =============================================================================

def slice_data(data_tuple, indices):
    """
    튜플 형태의 데이터셋(data, labels, rr...)을 인덱스에 맞춰 슬라이싱합니다.
    """
    sliced = []
    for item in data_tuple:
        if isinstance(item, np.ndarray) or isinstance(item, torch.Tensor):
            sliced.append(item[indices])
        else:
            # 리스트나 다른 타입인 경우
            sliced.append([item[i] for i in indices])
    return tuple(sliced)

def aggregate_metrics(fold_results):
    """
    Fold별 결과 리스트를 받아 평균(Mean)과 표준편차(Std)를 계산합니다. (Training 단계용)
    """
    agg = {}
    metrics_keys = fold_results[0].keys()
    
    for key in metrics_keys:
        values = [res[key] for res in fold_results]
        
        # 수치형 데이터만 계산
        if isinstance(values[0], (int, float, np.number)):
            agg[f"{key}_mean"] = np.mean(values)
            agg[f"{key}_std"] = np.std(values)
            # 원본 키에 평균값 저장 (엑셀 호환성 위함)
            agg[key] = np.mean(values)
        elif isinstance(values[0], np.ndarray):
            # Confusion Matrix 등 배열인 경우 합산/평균
            agg[key] = np.mean(values, axis=0)
            
    return agg


# =============================================================================
# Evaluation Helper Functions
# =============================================================================

def evaluate_model_cm(model, data_loader, device):
    """
    모델을 평가하여 Confusion Matrix를 반환합니다.
    """
    model.eval()
    all_preds = []
    all_labels = []

    with torch.no_grad():
        for batch in data_loader:
            # Unpack batch (Dataset에 따라 다름)
            # daeac/opt3: x, label, rr, ...
            # standard: x, label
            if len(batch) >= 2:
                inputs = batch[0].to(device)
                labels = batch[1].to(device)
                
                # RR interval 처리 (있으면)
                rr_features = None
                if len(batch) >= 3:
                     rr_features = batch[2].to(device)

                # Forward
                if rr_features is not None:
                     outputs = model(inputs, rr_features)
                else:
                     outputs = model(inputs)
                
                _, preds = torch.max(outputs, 1)
                
                all_preds.extend(preds.cpu().numpy())
                all_labels.extend(labels.cpu().numpy())

    # Calculate Confusion Matrix
    cm = confusion_matrix(all_labels, all_preds, labels=range(len(CLASSES)))
    return cm

def get_metrics_from_cm(cm):
    """Confusion Matrix로부터 모든 평가지표 계산"""
    support = cm.sum(axis=1)
    total = support.sum()
    sens, spec, prec, f1s, accs = [], [], [], [], []
    
    for i in range(len(CLASSES)):
        tp = cm[i, i]
        fp = cm[:, i].sum() - tp
        fn = cm[i, :].sum() - tp
        tn = cm.sum() - tp - fp - fn
        
        s = tp / (tp + fn) if (tp + fn) > 0 else 0
        sp = tn / (tn + fp) if (tn + fp) > 0 else 0
        p = tp / (tp + fp) if (tp + fp) > 0 else 0
        f = 2 * (p * s) / (p + s) if (p + s) > 0 else 0
        a = (tp + tn) / total
        
        sens.append(s); spec.append(sp); prec.append(p); f1s.append(f); accs.append(a)

    return {
        'macro': [np.mean(accs), np.mean(sens), np.mean(spec), np.mean(prec), np.mean(f1s)],
        'weighted': [
            np.trace(cm) / total, # accuracy
            np.sum(np.array(sens) * support) / total, # recall
            np.sum(np.array(spec) * support) / total,
            np.sum(np.array(prec) * support) / total,
            np.sum(np.array(f1s) * support) / total
        ],
        'per_class': {'acc': accs, 'recall': sens, 'spec': spec, 'prec': prec, 'f1': f1s}
    }


# =============================================================================
# Fold 실행 함수 (Training)
# =============================================================================

def run_single_fold(fold_idx, train_idx, test_idx, full_dataset_tuple, DatasetClass, 
                    exp_config, exp_name, device):
    """
    단일 Fold에 대한 학습 및 평가 수행
    """
    print(f"\n  [Fold {fold_idx+1}/{N_FOLDS}] Processing...")
    
    # 1. 데이터 분할 (Train/Test)
    train_data_split = slice_data(full_dataset_tuple, train_idx)
    test_data_split = slice_data(full_dataset_tuple, test_idx)
    
    # 2. Train을 다시 Train(90%)/Valid(10%)로 분할 (Early Stopping용)
    train_labels = train_data_split[1]
    
    tr_sub_idx, val_sub_idx = train_test_split(
        np.arange(len(train_labels)), 
        test_size=0.1, 
        stratify=train_labels, 
        random_state=config.SEED
    )
    
    real_train_data = slice_data(train_data_split, tr_sub_idx)
    real_valid_data = slice_data(train_data_split, val_sub_idx)
    
    # 3. Dataset & DataLoader 생성
    train_ds = DatasetClass(*real_train_data)
    valid_ds = DatasetClass(*real_valid_data)
    test_ds = DatasetClass(*test_data_split)
    
    train_loader = DataLoader(train_ds, batch_size=config.BATCH_SIZE, shuffle=True, num_workers=config.NUM_WORKERS, pin_memory=True)
    valid_loader = DataLoader(valid_ds, batch_size=config.BATCH_SIZE, shuffle=False, num_workers=config.NUM_WORKERS, pin_memory=True)
    test_loader = DataLoader(test_ds, batch_size=config.BATCH_SIZE, shuffle=False, num_workers=config.NUM_WORKERS, pin_memory=True)
    
    # 4. 모델 설정
    model_config = copy.deepcopy(config.MODEL_CONFIG)
    
    # Fusion/Opt3 설정 적용
    fusion_type = exp_config.get('fusion_type')
    is_opt3 = exp_config.get('use_opt3', False)
    
    if not is_opt3 and fusion_type is not None and fusion_type.lower() != 'none':
        model_config.update({
            "lead": FUSION_EXP_CONFIG["lead"],
            "rr_dim": FUSION_EXP_CONFIG["rr_dim"],
        })
    model_config.update(exp_config)
    
    # 5. 모델 초기화
    set_seed(config.SEED + fold_idx) 
    ModelClass = MODEL_REGISTRY.get(config.MODEL_NAME)
    model = ModelClass(**model_config)
    
    # 6. Loss & Optimizer
    class_weights = None
    if config.LOSS_CONFIG.get('use_class_weights', False):
        labels_np = real_train_data[1]
        counts = np.bincount(labels_np.flatten().astype(np.int64), minlength=config.NUM_CLASSES)
        weights = 1.0 / (counts + 1e-8)
        weights = weights / weights.sum() * config.NUM_CLASSES
        class_weights = torch.tensor(weights, dtype=torch.float32).to(device)

    LossClass = LOSS_REGISTRY.get(config.LOSS_NAME)
    criterion = LossClass(weight=class_weights)
    
    optimizer = build_optimizer(model.parameters(), config.OPTIMIZER_NAME, **config.OPTIMIZER_CONFIG)
    scheduler = build_scheduler(optimizer, config.SCHEDULER_NAME, **config.SCHEDULER_CONFIG)
    
    # 7. Trainer 실행
    fold_exp_dir = os.path.join(OUTPUT_DIR, exp_name, f"fold_{fold_idx+1}")
    os.makedirs(fold_exp_dir, exist_ok=True)
    
    trainer = Trainer(
        model=model,
        train_loader=train_loader,
        valid_loader=valid_loader,
        criterion=criterion,
        optimizer=optimizer,
        scheduler=scheduler,
        device=device,
        exp_dir=fold_exp_dir,
        class_names=config.CLASSES,
    )
    
    trainer.fit(epochs=EXP_EPOCHS)
    
    # 8. 평가 (Best Model 로드)
    trainer.load_best_model("macro_auprc")
    test_metrics = trainer.evaluate(test_loader)
    
    return test_metrics


# =============================================================================
# 실험 목록 생성
# =============================================================================

def generate_experiments():
    experiments = []
    for fusion_type in EXPERIMENT_GRID["fusion_type"]:
        if fusion_type is None:
            # Baseline: opt2
            exp_config = {"fusion_type": None, "lead": 3, "rr_dim": 2}
            exp_name = "baseline_opt2"
            experiments.append((exp_config, exp_name))
        elif fusion_type == "mhca":
            for num_heads in EXPERIMENT_GRID["fusion_num_heads"]:
                # Opt1 style
                exp_config = {"fusion_type": fusion_type, "fusion_num_heads": num_heads}
                exp_name = f"mhca_h{num_heads}"
                experiments.append((exp_config, exp_name))
                # Opt3 style
                exp_config_opt3 = {
                    "fusion_type": fusion_type, "fusion_num_heads": num_heads,
                    "lead": 3, "rr_dim": 7, "use_opt3": True
                }
                exp_name_opt3 = f"opt3_mhca_h{num_heads}"
                experiments.append((exp_config_opt3, exp_name_opt3))
        else:
            # Opt1 style
            exp_config = {"fusion_type": fusion_type}
            exp_name = f"fusion_{fusion_type}"
            experiments.append((exp_config, exp_name))
            # Opt3 style
            exp_config_opt3 = {
                "fusion_type": fusion_type,
                "lead": 3, "rr_dim": 7, "use_opt3": True
            }
            exp_name_opt3 = f"opt3_{fusion_type}"
            experiments.append((exp_config_opt3, exp_name_opt3))
    return experiments


# =============================================================================
# 메인 함수
# =============================================================================

def main():
    device = config.get_device()
    print(f"\n{'='*60}")
    print("AUTOMATED EXPERIMENT (INTRA-PATIENT / 5-FOLD CV)")
    print(f"{'='*60}")
    print(f"Device: {device}")
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    global OUTPUT_DIR
    OUTPUT_DIR = os.path.join(OUTPUT_DIR, timestamp)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    # -----------------------------------------------------------
    # 1. 데이터 로드 (전체 데이터를 한 번에 로드)
    # -----------------------------------------------------------
    print("\n[1/4] Loading ALL Data for Intra-patient split...")
    
    # Opt2 (DAEAC) 전체 로드
    print("  Loading Full opt2 (DAEAC) data...")
    full_data_opt2 = load_or_extract_data(
        record_list=ALL_RECORDS, 
        base_path=config.DATA_PATH, 
        valid_leads=config.VALID_LEADS, 
        out_len=config.SIGNAL_LENGTH, 
        split_name="Full_Opt2", 
        extraction_style="daeac"
    )
    
    # Opt1 (Standard) 전체 로드
    print("  Loading Full opt1 (Standard) data...")
    full_data_opt1 = load_or_extract_data(
        record_list=ALL_RECORDS, 
        base_path=config.DATA_PATH, 
        valid_leads=config.VALID_LEADS, 
        out_len=config.SIGNAL_LENGTH, 
        split_name="Full_Opt1", 
        extraction_style="default"
    )
    
    # Opt3 (Combined) 생성
    print("  Creating Full opt3 data...")
    def create_opt3_full(opt2_data, opt1_data):
        opt2_size = len(opt2_data[0])
        opt1_size = len(opt1_data[0])
        if opt2_size != opt1_size:
            print(f"  WARNING: Size mismatch! opt2={opt2_size}, opt1={opt1_size}")
            min_size = min(opt2_size, opt1_size)
            opt2_data = tuple(item[:min_size] if hasattr(item, '__getitem__') else item for item in opt2_data)
            opt1_data = tuple(item[:min_size] if hasattr(item, '__getitem__') else item for item in opt1_data)

        return (opt2_data[0], opt2_data[1], opt2_data[2], opt1_data[2], opt2_data[3], opt2_data[4])

    full_data_opt3 = create_opt3_full(full_data_opt2, full_data_opt1)
    
    # Dataset Registry Key 가져오기
    DatasetClass_opt1 = DATASET_REGISTRY.get("ecg_standard")
    DatasetClass_opt2 = DATASET_REGISTRY.get("daeac")
    DatasetClass_opt3 = DATASET_REGISTRY.get("opt3")
    
    full_datasets_map = {
        "opt1": (full_data_opt1, DatasetClass_opt1),
        "opt2": (full_data_opt2, DatasetClass_opt2),
        "opt3": (full_data_opt3, DatasetClass_opt3)
    }
    
    print(f"  Total Samples: {len(full_data_opt2[0])}")
    
    # -----------------------------------------------------------
    # 2. 실험 준비
    # -----------------------------------------------------------
    experiments = generate_experiments()
    print(f"\n[2/4] Running {len(experiments)} experiments with {N_FOLDS}-Fold CV...")
    
    template_path = "./model_fusion.xlsx"
    excel_writer = None
    cumulative_writer = None
    
    if os.path.exists(template_path):
        excel_path = os.path.join(OUTPUT_DIR, "autoexp_intra_results.xlsx")
        excel_writer = ExcelResultWriter(template_path, excel_path, classes=config.CLASSES)
        cumulative_path = "./cumulative_intra_results.xlsx"
        cumulative_writer = CumulativeExcelWriter(template_path, cumulative_path, classes=config.CLASSES)

    # -----------------------------------------------------------
    # 3. 실험 실행 Loop (Training)
    # -----------------------------------------------------------
    total_start = time.time()
    
    # Evaluation에서 사용할 정보 저장용
    experiment_info_list = [] # (exp_name, exp_config, dataset_key)

    for i, (exp_config, exp_name) in enumerate(experiments, 1):
        print(f"\n{'='*80}")
        print(f"[{i}/{len(experiments)}] Experiment: {exp_name}")
        print(f"Config: {exp_config}")
        print(f"{'='*80}")
        
        # 데이터셋 선택
        if exp_config.get('use_opt3', False):
            target_key = "opt3"
        elif exp_config.get('fusion_type') is not None:
            target_key = "opt1"
        else:
            target_key = "opt2"
        
        experiment_info_list.append({
            "exp_name": exp_name,
            "exp_config": exp_config,
            "dataset_key": target_key
        })
            
        full_data_tuple, DatasetClass = full_datasets_map[target_key]
        all_labels = full_data_tuple[1] 
        
        skf = StratifiedKFold(n_splits=N_FOLDS, shuffle=True, random_state=config.SEED)
        fold_metrics_list = []
        
        try:
            for fold_idx, (train_idx, test_idx) in enumerate(skf.split(np.zeros(len(all_labels)), all_labels)):
                fold_res = run_single_fold(
                    fold_idx, train_idx, test_idx, 
                    full_data_tuple, DatasetClass, 
                    exp_config, exp_name, device
                )
                fold_metrics_list.append(fold_res)
                print(f"  -> Fold {fold_idx+1} Result: Acc={fold_res['acc']:.4f}, F1={fold_res['macro_f1']:.4f}, AUPRC={fold_res['macro_auprc']:.4f}")

            aggregated_results = aggregate_metrics(fold_metrics_list)
            
            print(f"\n  [Experiment Complete] {exp_name}")
            print(f"  Mean Acc: {aggregated_results['acc_mean']:.4f} ± {aggregated_results['acc_std']:.4f}")
            
            if excel_writer:
                excel_writer.write_metrics(exp_name, aggregated_results, "intra_mean")
            
            if cumulative_writer:
                cumulative_writer.append_result(exp_name, aggregated_results, exp_config, "intra_mean")

        except Exception as e:
            print(f"  ERROR in {exp_name}: {e}")
            import traceback
            traceback.print_exc()

    # -----------------------------------------------------------
    # 4. 종합 평가 (Evaluation & Excel Report)
    # -----------------------------------------------------------
    print(f"\n{'='*60}")
    print("[3/4] STARTING COMPREHENSIVE EVALUATION (CM Analysis)")
    print(f"{'='*60}")

    CM_DATA = {}

    for info in experiment_info_list:
        exp_name = info['exp_name']
        exp_config = info['exp_config']
        dataset_key = info['dataset_key']
        
        print(f"\n  Evaluating: {exp_name}")
        
        full_data_tuple, DatasetClass = full_datasets_map[dataset_key]
        all_labels = full_data_tuple[1]
        
        CM_DATA[exp_name] = {}
        
        # Training과 동일한 Split 재현
        skf = StratifiedKFold(n_splits=N_FOLDS, shuffle=True, random_state=config.SEED)

        for fold_idx, (train_idx, test_idx) in enumerate(skf.split(np.zeros(len(all_labels)), all_labels)):
            fold_num = fold_idx + 1
            
            # Checkpoint 경로 (Trainer가 저장한 경로)
            ckpt_path = os.path.join(OUTPUT_DIR, exp_name, f"fold_{fold_num}", "best_models", "best_macro_auprc.pth")
            
            if not os.path.exists(ckpt_path):
                print(f"    [Fold {fold_num}] Checkpoint not found: {ckpt_path}")
                continue
                
            # 모델 설정 및 로드
            model_config = copy.deepcopy(config.MODEL_CONFIG)
            
            is_opt3 = exp_config.get('use_opt3', False)
            fusion_type = exp_config.get('fusion_type')

            if not is_opt3 and fusion_type is not None and fusion_type.lower() != 'none':
                model_config.update({
                    "lead": FUSION_EXP_CONFIG["lead"],
                    "rr_dim": FUSION_EXP_CONFIG["rr_dim"],
                })
            model_config.update(exp_config)
            
            set_seed(config.SEED + fold_idx)
            ModelClass = MODEL_REGISTRY.get(config.MODEL_NAME)
            model = ModelClass(**model_config)
            
            checkpoint = torch.load(ckpt_path, map_location=device, weights_only=False)
            model.load_state_dict(checkpoint['model_state_dict'])
            model.to(device)
            
            # Test Loader 준비
            test_data_split = slice_data(full_data_tuple, test_idx)
            test_ds = DatasetClass(*test_data_split)
            test_loader = DataLoader(test_ds, batch_size=config.BATCH_SIZE, shuffle=False, num_workers=config.NUM_WORKERS, pin_memory=True)
            
            # Confusion Matrix 계산
            cm = evaluate_model_cm(model, test_loader, device)
            CM_DATA[exp_name][fold_num] = cm
            
            print(f"    [Fold {fold_num}] CM Generated. Acc: {np.trace(cm)/cm.sum():.4f}")

    # -----------------------------------------------------------
    # 5. Excel 저장 (Mean ± Std)
    # -----------------------------------------------------------
    print("\n[4/4] Saving Final Report to Excel...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Performance"

    # Header
    ws.cell(1, 1).value = "Experiment"
    header = ["Acc", "Recall", "Spec", "Prec", "F1"]
    for i, group in enumerate(["Macro", "Weighted"] + CLASSES):
        for j, h in enumerate(header):
            ws.cell(1, 2 + i*5 + j).value = f"{group}_{h}"

    # Write Data
    for row_idx, model_key in enumerate(sorted(CM_DATA.keys()), 2):
        if len(CM_DATA[model_key]) < N_FOLDS:
            print(f"  Warning: {model_key} has incomplete folds")
            continue

        folds = [get_metrics_from_cm(CM_DATA[model_key][f]) for f in sorted(CM_DATA[model_key].keys())]
        
        ws.cell(row_idx, 1).value = model_key

        def write_mean_std(start_col, data_list):
            means = np.mean(data_list, axis=0)
            stds = np.std(data_list, axis=0)
            for i in range(5):
                ws.cell(row_idx, start_col + i).value = f"{means[i]:.4f}±{stds[i]:.4f}"

        # Macro (Col 2-6)
        write_mean_std(2, [f['macro'] for f in folds])
        # Weighted (Col 7-11)
        write_mean_std(7, [f['weighted'] for f in folds])
        # Per-class (Col 12+)
        for c_idx in range(len(CLASSES)):
            base_col = 12 + c_idx * 5
            class_data = []
            for f in folds:
                p = f['per_class']
                class_data.append([p['acc'][c_idx], p['recall'][c_idx], p['spec'][c_idx], p['prec'][c_idx], p['f1'][c_idx]])
            write_mean_std(base_col, class_data)

    final_report_path = os.path.join(OUTPUT_DIR, "ECG_CM_Analysis_Result.xlsx")
    wb.save(final_report_path)
    
    total_time = (time.time() - total_start) / 60
    print(f"\n{'='*60}")
    print("ALL TASKS COMPLETED SUCCESSFULLY")
    print(f"Total time: {total_time:.1f} min")
    print(f"Final Report: {final_report_path}")
    print(f"{'='*60}")

if __name__ == '__main__':
    main()