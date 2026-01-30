"""
Evaluation Script for Intra-Patient Experiments

autoexp_results_intra_202602/20260129_213034 폴더 내 실험들을 평가하고
Confusion Matrix 기반 metrics를 계산하여 Excel로 저장합니다.
"""

import os
import copy
import numpy as np
import torch
from torch.utils.data import DataLoader
from sklearn.model_selection import StratifiedKFold
from sklearn.metrics import confusion_matrix
from openpyxl import Workbook

import config
from src.registry import MODEL_REGISTRY, DATASET_REGISTRY
from src import models, datasets
from utils import set_seed, load_or_extract_data


# =============================================================================
# 설정
# =============================================================================
RESULTS_DIR = "/home/work/Ryuha/autoexp_results_intra_202602/20260129_213034"
N_FOLDS = 5
CLASSES = ["N", "S", "V", "F"]

# 실험 설정 (main_autoexp_intra.py의 generate_experiments 참조)
EXPERIMENTS = {
    "baseline_opt2": {
        "fusion_type": None,
        "lead": 3,
        "rr_dim": 2,
        "dataset_key": "opt2"
    },
    "fusion_concat": {
        "fusion_type": "concat",
        "lead": 1,
        "rr_dim": 7,
        "dataset_key": "opt1"
    },
    "opt3_concat": {
        "fusion_type": "concat",
        "lead": 3,
        "rr_dim": 7,
        "use_opt3": True,
        "dataset_key": "opt3"
    },
    "mhca_h1": {
        "fusion_type": "mhca",
        "fusion_num_heads": 1,
        "lead": 1,
        "rr_dim": 7,
        "dataset_key": "opt1"
    },
    "opt3_mhca_h1": {
        "fusion_type": "mhca",
        "fusion_num_heads": 1,
        "lead": 3,
        "rr_dim": 7,
        "use_opt3": True,
        "dataset_key": "opt3"
    },
}

ALL_RECORDS = config.DS1_TRAIN + config.DS1_VALID + config.DS2_TEST


# =============================================================================
# 유틸리티 함수
# =============================================================================
def slice_data(data_tuple, indices):
    """튜플 형태의 데이터셋을 인덱스에 맞춰 슬라이싱"""
    sliced = []
    for item in data_tuple:
        if isinstance(item, np.ndarray) or isinstance(item, torch.Tensor):
            sliced.append(item[indices])
        else:
            sliced.append([item[i] for i in indices])
    return tuple(sliced)


def get_metrics_from_cm(cm):
    """Confusion Matrix로부터 모든 평가지표 계산"""
    support = cm.sum(axis=1)
    total = support.sum()
    sens, spec, prec, f1s, accs = [], [], [], [], []

    for i in range(len(CLASSES)):
        tp = cm[i, i]
        fp = cm[:, i].sum() - tp
        fn = cm[i, :].sum() - tp
        tn = cm.sum() - tp - fp - fn

        s = tp / (tp + fn) if (tp + fn) > 0 else 0
        sp = tn / (tn + fp) if (tn + fp) > 0 else 0
        p = tp / (tp + fp) if (tp + fp) > 0 else 0
        f = 2 * (p * s) / (p + s) if (p + s) > 0 else 0
        a = (tp + tn) / total

        sens.append(s)
        spec.append(sp)
        prec.append(p)
        f1s.append(f)
        accs.append(a)

    return {
        'macro': [np.mean(accs), np.mean(sens), np.mean(spec), np.mean(prec), np.mean(f1s)],
        'weighted': [
            np.trace(cm) / total,  # accuracy
            np.sum(np.array(sens) * support) / total,  # recall
            np.sum(np.array(spec) * support) / total,
            np.sum(np.array(prec) * support) / total,
            np.sum(np.array(f1s) * support) / total
        ],
        'per_class': {'acc': accs, 'recall': sens, 'spec': spec, 'prec': prec, 'f1': f1s}
    }


def evaluate_model(model, loader, device):
    """모델 평가 후 confusion matrix 반환"""
    model.eval()
    all_labels = []
    all_preds = []

    with torch.no_grad():
        for batch in loader:
            ecg, labels, rr_features, *_ = batch
            ecg = ecg.to(device)
            rr_features = rr_features.to(device)

            logits, _ = model(ecg, rr_features)
            preds = torch.argmax(logits, dim=1)

            all_labels.append(labels.numpy())
            all_preds.append(preds.cpu().numpy())

    y_true = np.concatenate(all_labels)
    y_pred = np.concatenate(all_preds)

    cm = confusion_matrix(y_true, y_pred, labels=range(len(CLASSES)))
    return cm


# =============================================================================
# 메인 함수
# =============================================================================
def main():
    device = config.get_device()
    print(f"\n{'='*60}")
    print("EVALUATION SCRIPT - Intra-Patient Experiments")
    print(f"{'='*60}")
    print(f"Device: {device}")
    print(f"Results Dir: {RESULTS_DIR}")

    # -----------------------------------------------------------
    # 1. 데이터 로드
    # -----------------------------------------------------------
    print("\n[1/3] Loading ALL Data...")

    # Opt2 (DAEAC) 전체 로드
    print("  Loading opt2 (DAEAC) data...")
    full_data_opt2 = load_or_extract_data(
        record_list=ALL_RECORDS,
        base_path=config.DATA_PATH,
        valid_leads=config.VALID_LEADS,
        out_len=config.SIGNAL_LENGTH,
        split_name="Full_Opt2",
        extraction_style="daeac"
    )

    # Opt1 (Standard) 전체 로드
    print("  Loading opt1 (Standard) data...")
    full_data_opt1 = load_or_extract_data(
        record_list=ALL_RECORDS,
        base_path=config.DATA_PATH,
        valid_leads=config.VALID_LEADS,
        out_len=config.SIGNAL_LENGTH,
        split_name="Full_Opt1",
        extraction_style="default"
    )

    # Opt3 (Combined) 생성 - Training과 동일한 로직 사용
    print("  Creating opt3 data...")
    opt2_size = len(full_data_opt2[0])
    opt1_size = len(full_data_opt1[0])
    print(f"  opt2 size: {opt2_size}, opt1 size: {opt1_size}")

    # opt3용 데이터만 truncate (opt1, opt2는 원본 유지)
    if opt2_size != opt1_size:
        min_size = min(opt2_size, opt1_size)
        print(f"  opt3 will use truncated size: {min_size}")
        opt2_for_opt3 = tuple(item[:min_size] if hasattr(item, '__getitem__') else item for item in full_data_opt2)
        opt1_for_opt3 = tuple(item[:min_size] if hasattr(item, '__getitem__') else item for item in full_data_opt1)
    else:
        opt2_for_opt3 = full_data_opt2
        opt1_for_opt3 = full_data_opt1

    full_data_opt3 = (opt2_for_opt3[0], opt2_for_opt3[1], opt2_for_opt3[2],
                      opt1_for_opt3[2], opt2_for_opt3[3], opt2_for_opt3[4])

    # Dataset Classes
    DatasetClass_opt1 = DATASET_REGISTRY.get("ecg_standard")
    DatasetClass_opt2 = DATASET_REGISTRY.get("daeac")
    DatasetClass_opt3 = DATASET_REGISTRY.get("opt3")

    full_datasets_map = {
        "opt1": (full_data_opt1, DatasetClass_opt1),  # 원본 유지
        "opt2": (full_data_opt2, DatasetClass_opt2),  # 원본 유지
        "opt3": (full_data_opt3, DatasetClass_opt3)   # truncated
    }

    print(f"  opt1 Samples: {len(full_data_opt1[0])}")
    print(f"  opt2 Samples: {len(full_data_opt2[0])}")
    print(f"  opt3 Samples: {len(full_data_opt3[0])}")

    # -----------------------------------------------------------
    # 2. 실험별 평가
    # -----------------------------------------------------------
    print("\n[2/3] Evaluating experiments...")

    # 결과 저장: {exp_name: {fold_idx: confusion_matrix}}
    CM_DATA = {}

    # 사용 가능한 실험 폴더 확인
    available_exps = [d for d in os.listdir(RESULTS_DIR)
                      if os.path.isdir(os.path.join(RESULTS_DIR, d)) and d in EXPERIMENTS]

    print(f"  Found experiments: {available_exps}")

    for exp_name in available_exps:
        print(f"\n  Evaluating: {exp_name}")
        exp_config = EXPERIMENTS[exp_name]
        dataset_key = exp_config["dataset_key"]

        full_data_tuple, DatasetClass = full_datasets_map[dataset_key]
        all_labels = full_data_tuple[1]

        CM_DATA[exp_name] = {}

        # K-Fold (동일한 random_state로 동일한 split 재현)
        skf = StratifiedKFold(n_splits=N_FOLDS, shuffle=True, random_state=config.SEED)

        for fold_idx, (train_idx, test_idx) in enumerate(skf.split(np.zeros(len(all_labels)), all_labels)):
            fold_num = fold_idx + 1

            # 체크포인트 경로
            ckpt_path = os.path.join(RESULTS_DIR, exp_name, f"fold_{fold_num}",
                                     "best_models", "best_macro_auprc.pth")

            if not os.path.exists(ckpt_path):
                print(f"    [Fold {fold_num}] Checkpoint not found: {ckpt_path}")
                continue

            # 모델 설정
            model_config = copy.deepcopy(config.MODEL_CONFIG)
            model_config.update({
                "lead": exp_config.get("lead", 1),
                "rr_dim": exp_config.get("rr_dim", 7),
                "fusion_type": exp_config.get("fusion_type"),
                "fusion_num_heads": exp_config.get("fusion_num_heads", 1),
            })

            # 모델 생성 및 로드
            set_seed(config.SEED + fold_idx)
            ModelClass = MODEL_REGISTRY.get(config.MODEL_NAME)
            model = ModelClass(**model_config)

            checkpoint = torch.load(ckpt_path, map_location=device, weights_only=False)
            model.load_state_dict(checkpoint['model_state_dict'])
            model.to(device)

            # 테스트 데이터 준비
            test_data_split = slice_data(full_data_tuple, test_idx)
            test_ds = DatasetClass(*test_data_split)
            test_loader = DataLoader(test_ds, batch_size=config.BATCH_SIZE,
                                     shuffle=False, num_workers=config.NUM_WORKERS, pin_memory=True)

            # 평가
            cm = evaluate_model(model, test_loader, device)
            CM_DATA[exp_name][fold_num] = cm

            print(f"    [Fold {fold_num}] Done - Acc: {np.trace(cm)/cm.sum():.4f}")

    # -----------------------------------------------------------
    # 3. Excel 저장
    # -----------------------------------------------------------
    print("\n[3/3] Saving results to Excel...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Performance"

    # Header 설정
    ws.cell(1, 1).value = "Experiment"
    header = ["Acc", "Recall", "Spec", "Prec", "F1"]
    for i, group in enumerate(["Macro", "Weighted"] + CLASSES):
        for j, h in enumerate(header):
            ws.cell(1, 2 + i*5 + j).value = f"{group}_{h}"

    # 모델별 데이터 처리
    for row_idx, model_key in enumerate(sorted(CM_DATA.keys()), 2):
        if len(CM_DATA[model_key]) < N_FOLDS:
            print(f"  Warning: {model_key} has only {len(CM_DATA[model_key])} folds")
            continue

        folds = [get_metrics_from_cm(CM_DATA[model_key][f]) for f in range(1, 6)]

        ws.cell(row_idx, 1).value = model_key

        def write_mean_std(start_col, data_list):
            means = np.mean(data_list, axis=0)
            stds = np.std(data_list, axis=0)
            for i in range(5):
                ws.cell(row_idx, start_col + i).value = f"{means[i]:.4f}±{stds[i]:.4f}"

        # 1. Macro (Col 2-6)
        write_mean_std(2, [f['macro'] for f in folds])

        # 2. Weighted (Col 7-11)
        write_mean_std(7, [f['weighted'] for f in folds])

        # 3. Per-class (Col 12+)
        for c_idx in range(len(CLASSES)):
            base_col = 12 + c_idx * 5
            class_data = []
            for f in folds:
                p = f['per_class']
                class_data.append([p['acc'][c_idx], p['recall'][c_idx], p['spec'][c_idx],
                                   p['prec'][c_idx], p['f1'][c_idx]])
            write_mean_std(base_col, class_data)

    save_path = os.path.join(RESULTS_DIR, "ECG_CM_Analysis_Result.xlsx")
    wb.save(save_path)
    print(f"\nExcel file saved to: {save_path}")

    print(f"\n{'='*60}")
    print("EVALUATION COMPLETE")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
