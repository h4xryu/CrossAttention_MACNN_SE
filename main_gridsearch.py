"""
main_gridsearch.py - MACNN_SE 모델 하이퍼파라미터 그리드 서치
MACNN_SE 클래스의 fusion_emb, fusion_expansion, num_heads, reduction,
aspp_bn, aspp_act, dilations 파라미터에 대한 그리드 서치

결과:
- model_gridsearch.xlsx 템플릿에 결과 저장 (Performance Metrics, Confusion 시트)
- Best AUROC, Best AUPRC 모델 각각 저장
"""

import os
import time
import copy
import shutil
from collections import Counter
from datetime import datetime
from itertools import product

import numpy as np
import pandas as pd
import torch
from torch.utils.data import DataLoader
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from logger import (
    TrainingLogger, print_epoch_header, print_per_class_metrics,
    print_epoch_stats, print_confidence_stats, print_epoch_time
)

from utils import set_seed, load_or_extract_data
from model import MACNN_SE
from dataloader import ECGDataset
from train import train_one_epoch, validate, save_model
from evaluate_module import evaluate, calculate_metrics, save_confusion_matrix

# =============================================================================
# 설정
# =============================================================================

DATA_PATH = './data/mit-bih-arrhythmia-database-1.0.0/'
OUTPUT_PATH = './gridsearch_results/'
TEMPLATE_PATH = './model_gridsearch.xlsx'
BATCH_SIZE = 1024 * 8
EPOCHS = 50
LR = 0.0001
WEIGHT_DECAY = 1e-3
SEED = 1234
POLY1_EPS = 0.0
POLY2_EPS = 0.0
CLASSES = ['N', 'S', 'V', 'F']

# RR Feature 설정
RR_FEATURE_OPTION = "opt3"
RR_FEATURE_DIMS = {"opt1": 7, "opt2": 38, "opt3": 7, "opt4": 7}

# ECG Parameters
VALID_LEADS = ['MLII', 'V1', 'V2', 'V4', 'V5']
OUT_LEN = 720

# 데이터 분할
DS1_TRAIN_SPLIT = [
    '101', '106', '108', '109', '112', '115', '116', '118', '119',
    '122', '201', '203', '209', '215', '223', '230', '208'
]
DS1_VALID_SPLIT = ['114', '124', '205', '207', '220']

DS2_TEST = [
    '100', '103', '105', '111', '113', '117', '121', '123', '200', '202',
    '210', '212', '213', '214', '219', '221', '222', '228', '231', '232',
    '233', '234'
]

# =============================================================================
# 그리드 서치 파라미터 - MACNN_SE 클래스 기준
# =============================================================================

GRID_PARAMS = {

    'fusion_emb': [64, 128],
    'fusion_expansion': [2, 4],
    'num_heads': [1],
    'reduction': [8, 16],
    'aspp_bn': [True],
    'aspp_act': [True],
    'dilations': [(1, 6, 12, 18), (1, 3, 6, 12)],
}

# =============================================================================
# 엑셀 결과 저장 클래스
# =============================================================================

class ExcelResultWriter:
    """엑셀 템플릿에 결과를 실시간으로 저장하는 클래스"""

    def __init__(self, template_path, output_path):
        self.template_path = template_path
        self.output_path = output_path
        self.current_row = 3  # 데이터 시작 행 (0-indexed, 헤더 3줄 이후)
        self.confusion_start_row = 0  # Confusion 시트 시작 행

        # 템플릿 복사
        shutil.copy(template_path, output_path)
        print(f"Template copied to: {output_path}")

    def write_metrics(self, exp_name: str, metrics: dict, best_type: str):
        """
        Performance Metrics 시트에 결과 작성

        Args:
            exp_name: 실험 이름 (예: "emb64_exp2_h1_r16@")
            metrics: calculate_metrics 결과
            best_type: 'auroc' 또는 'auprc'
        """
        wb = load_workbook(self.output_path)
        ws = wb['Performance Metrics']

        row = self.current_row + 1  # openpyxl은 1-indexed

        # 실험명 (best_type 포함)
        full_name = f"{exp_name}_{best_type}"
        ws.cell(row=row, column=1, value=full_name)

        # Macro metrics (columns 2-6)
        ws.cell(row=row, column=2, value=metrics['macro_accuracy'])
        ws.cell(row=row, column=3, value=metrics['macro_recall'])  # Sensitivity
        ws.cell(row=row, column=4, value=metrics['macro_specificity'])
        ws.cell(row=row, column=5, value=metrics['macro_prec'])
        ws.cell(row=row, column=6, value=metrics['macro_f1'])

        # Weighted metrics (columns 7-11)
        ws.cell(row=row, column=7, value=metrics['weighted_accuracy'])
        ws.cell(row=row, column=8, value=metrics['weighted_recall'])
        ws.cell(row=row, column=9, value=metrics['weighted_specificity'])
        ws.cell(row=row, column=10, value=metrics['weighted_prec'])
        ws.cell(row=row, column=11, value=metrics['weighted_f1'])

        # Per-class metrics (N, S, V, F)
        for i, cls in enumerate(CLASSES):
            base_col = 12 + i * 5
            ws.cell(row=row, column=base_col, value=metrics['per_class_accuracy'][i])
            ws.cell(row=row, column=base_col + 1, value=metrics['per_class_recall'][i])
            ws.cell(row=row, column=base_col + 2, value=metrics['per_class_specificity'][i])
            ws.cell(row=row, column=base_col + 3, value=metrics['per_class_precision'][i])
            ws.cell(row=row, column=base_col + 4, value=metrics['per_class_f1'][i])

        wb.save(self.output_path)
        self.current_row += 1
        print(f"  Results written to row {row}: {full_name}")

    def write_confusion_matrix(self, exp_name: str, cm: np.ndarray, best_type: str):
        """
        Confusion 시트에 혼동 행렬 작성

        Args:
            exp_name: 실험 이름
            cm: 혼동 행렬 (4x4)
            best_type: 'auroc' 또는 'auprc'
        """
        wb = load_workbook(self.output_path)
        ws = wb['Confusion']

        # 새로운 confusion matrix 블록 시작 위치 찾기
        # 각 블록은 8행 (EXPNAME, 헤더2줄, 데이터4줄, 빈줄1개)
        block_start = self.confusion_start_row + 1  # openpyxl은 1-indexed

        full_name = f"{exp_name}_{best_type}"

        # EXPNAME 행
        ws.cell(row=block_start, column=1, value=full_name)

        # Predicted 헤더
        ws.cell(row=block_start + 1, column=3, value="Predicted")

        # 클래스 헤더
        for i, cls in enumerate(CLASSES):
            ws.cell(row=block_start + 2, column=3 + i, value=cls)

        # Actual 라벨 및 데이터
        ws.cell(row=block_start + 3, column=1, value="Actual")
        for i, cls in enumerate(CLASSES):
            ws.cell(row=block_start + 3 + i, column=2, value=cls)
            for j in range(4):
                ws.cell(row=block_start + 3 + i, column=3 + j, value=int(cm[i, j]))

        wb.save(self.output_path)
        self.confusion_start_row += 8  # 다음 블록 위치
        print(f"  Confusion matrix written for: {full_name}")


# =============================================================================
# 단일 실험 수행 함수
# =============================================================================

def run_single_experiment(params: dict, trial_num: int, total_trials: int,
                          device: torch.device, data_loaders: tuple,
                          excel_writer: ExcelResultWriter, exp_dir: str):
    """
    단일 그리드 서치 시행

    Args:
        params: 하이퍼파라미터 딕셔너리
        trial_num: 현재 시행 번호
        total_trials: 전체 시행 수
        device: torch device
        data_loaders: (train_loader, valid_loader, test_loader)
        excel_writer: 엑셀 결과 저장 객체
        exp_dir: 실험 디렉토리

    Returns:
        result_dict: 결과 딕셔너리
    """
    train_loader, valid_loader, test_loader = data_loaders

    # 실험 이름 생성
    exp_name = f"emb{params['fusion_emb']}_exp{params['fusion_expansion']}_h{params['num_heads']}_r{params['reduction']}"
    if params['dilations'] != (1, 6, 12, 18):
        exp_name += f"_d{''.join(map(str, params['dilations']))}"

    print(f"\n{'='*80}")
    print(f"Trial {trial_num}/{total_trials}: {exp_name}")
    print(f"  Params: {params}")
    print(f"{'='*80}")

    result_dict = {
        'exp_name': exp_name,
        'params': params,
        'best_valid_auroc': 0.0,
        'best_valid_auprc': 0.0,
        'test_auroc_metrics': None,
        'test_auprc_metrics': None,
        'status': 'error'
    }

    try:
        set_seed(SEED)

        # MACNN_SE 모델 생성 - 직접 클래스 사용
        model = MACNN_SE(
            reduction=params['reduction'],
            aspp_bn=params['aspp_bn'],
            aspp_act=params['aspp_act'],
            lead=1,
            p=0.0,
            dilations=params['dilations'],
            act_func='tanh',
            f_act_func='tanh',
            apply_residual=False,
            fusion_type='mhca',  # MHCA 사용 (그리드 서치에서는 고정)
            fusion_emb=params['fusion_emb'],
            fusion_expansion=params['fusion_expansion'],
            rr_dim=RR_FEATURE_DIMS[RR_FEATURE_OPTION],
            num_heads=params['num_heads']
        ).to(device)

        # 학습 설정
        optimizer = torch.optim.AdamW(model.parameters(), lr=LR, weight_decay=WEIGHT_DECAY)
        scheduler = torch.optim.lr_scheduler.StepLR(optimizer, step_size=20, gamma=0.5)
        logger = TrainingLogger(os.path.join(exp_dir, 'runs'))
        # Best 모델 추적
        best_auroc = {'value': 0.0, 'epoch': 0, 'state_dict': None}
        best_auprc = {'value': 0.0, 'epoch': 0, 'state_dict': None}

        # 학습 루프
        for epoch in range(1, EPOCHS + 1):
            current_lr = optimizer.param_groups[0]['lr']

            # Train
            train_loss, train_metrics, *p_t_train = train_one_epoch(
                model, train_loader, POLY1_EPS, POLY2_EPS, epoch, optimizer, device
            )

            print_epoch_stats(epoch, train_loss, train_metrics['acc'], current_lr, phase='Train')
            # print_per_class_metrics(train_metrics, CLASSES, phase='Train')
            # print_confidence_stats(*p_t_train, phase='Train')
            logger.log_epoch(epoch, train_loss, train_metrics, phase='train')
            logger.log_confidence(epoch, *p_t_train, phase='train')
            # Validation
            valid_loss, valid_metrics, *p_t_valid = validate(
                model, valid_loader, POLY1_EPS, POLY2_EPS, device
            )

            print_epoch_stats(epoch, valid_loss, valid_metrics['acc'], current_lr, phase='Valid')
            # print_per_class_metrics(valid_metrics, CLASSES, phase='Valid')
            # print_confidence_stats(*p_t_valid, phase='Valid')
            logger.log_epoch(epoch, valid_loss, valid_metrics, phase='valid')
            logger.log_confidence(epoch, *p_t_valid, phase='valid')
            # Best AUROC 체크
            if valid_metrics['macro_auroc'] > best_auroc['value']:
                best_auroc = {
                    'value': valid_metrics['macro_auroc'],
                    'epoch': epoch,
                    'state_dict': copy.deepcopy(model.state_dict())
                }
                print(f"  ★ [BEST AUROC] {best_auroc['value']:.4f}")

            # Best AUPRC 체크
            if valid_metrics['macro_auprc'] > best_auprc['value']:
                best_auprc = {
                    'value': valid_metrics['macro_auprc'],
                    'epoch': epoch,
                    'state_dict': copy.deepcopy(model.state_dict())
                }
                print(f"  ★ [BEST AUPRC] {best_auprc['value']:.4f}")

            scheduler.step()
            print("=" * 120 + "\n")
        logger.close()    
        # =====================================================================
        # Best AUROC 모델로 테스트
        # =====================================================================
        print(f"\n  Testing Best AUROC model (epoch {best_auroc['epoch']})...")
        model.load_state_dict(best_auroc['state_dict'])
        model.eval()

        y_pred_auroc, y_true_auroc, _ = evaluate(model, test_loader, device)
        metrics_auroc = calculate_metrics(np.array(y_true_auroc), np.array(y_pred_auroc))

        # 모델 저장
        auroc_path = os.path.join(exp_dir, 'best_weights', f'{exp_name}_best_auroc.pth')
        torch.save({
            'model_state_dict': best_auroc['state_dict'],
            'epoch': best_auroc['epoch'],
            'params': params,
            'auroc': best_auroc['value']
        }, auroc_path)

        # Confusion matrix 이미지 저장
        cm_auroc_path = os.path.join(exp_dir, 'confusion_matrices', f'{exp_name}_auroc.png')
        save_confusion_matrix(metrics_auroc['confusion_matrix'], CLASSES, cm_auroc_path)

        # 엑셀에 결과 작성
        excel_writer.write_metrics(exp_name, metrics_auroc, 'auroc')
        excel_writer.write_confusion_matrix(exp_name, metrics_auroc['confusion_matrix'], 'auroc')

        print(f"  AUROC Test - Macro F1: {metrics_auroc['macro_f1']:.4f}, "
              f"Acc: {metrics_auroc['overall_accuracy']:.4f}")

        # =====================================================================
        # Best AUPRC 모델로 테스트
        # =====================================================================
        print(f"\n  Testing Best AUPRC model (epoch {best_auprc['epoch']})...")
        model.load_state_dict(best_auprc['state_dict'])
        model.eval()

        y_pred_auprc, y_true_auprc, _ = evaluate(model, test_loader, device)
        metrics_auprc = calculate_metrics(np.array(y_true_auprc), np.array(y_pred_auprc))

        # 모델 저장
        auprc_path = os.path.join(exp_dir, 'best_weights', f'{exp_name}_best_auprc.pth')
        torch.save({
            'model_state_dict': best_auprc['state_dict'],
            'epoch': best_auprc['epoch'],
            'params': params,
            'auprc': best_auprc['value']
        }, auprc_path)

        # Confusion matrix 이미지 저장
        cm_auprc_path = os.path.join(exp_dir, 'confusion_matrices', f'{exp_name}_auprc.png')
        save_confusion_matrix(metrics_auprc['confusion_matrix'], CLASSES, cm_auprc_path)

        # 엑셀에 결과 작성
        excel_writer.write_metrics(exp_name, metrics_auprc, 'auprc')
        excel_writer.write_confusion_matrix(exp_name, metrics_auprc['confusion_matrix'], 'auprc')

        print(f"  AUPRC Test - Macro F1: {metrics_auprc['macro_f1']:.4f}, "
              f"Acc: {metrics_auprc['overall_accuracy']:.4f}")

        # 결과 저장
        result_dict['best_valid_auroc'] = best_auroc['value']
        result_dict['best_valid_auprc'] = best_auprc['value']
        result_dict['test_auroc_metrics'] = metrics_auroc
        result_dict['test_auprc_metrics'] = metrics_auprc
        result_dict['status'] = 'success'

    except Exception as e:
        print(f"  Error: {str(e)}")
        import traceback
        traceback.print_exc()
        result_dict['status'] = 'error'
   
    return result_dict


# =============================================================================
# 메인 그리드 서치 함수
# =============================================================================

def run_gridsearch():
    """전체 그리드 서치 실행"""
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

    print(f"\n{'='*80}")
    print(f"MACNN_SE Model Hyperparameter Grid Search")
    print(f"{'='*80}")
    print(f"Device: {device}")
    if torch.cuda.is_available():
        print(f"GPU: {torch.cuda.get_device_name(0)}")

    # 실험 디렉토리 생성
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    exp_dir = os.path.join(OUTPUT_PATH, f'gridsearch_{timestamp}')
    os.makedirs(exp_dir, exist_ok=True)
    os.makedirs(os.path.join(exp_dir, 'best_weights'), exist_ok=True)
    os.makedirs(os.path.join(exp_dir, 'confusion_matrices'), exist_ok=True)

    # 엑셀 결과 저장 객체 생성
    output_xlsx = os.path.join(exp_dir, f'gridsearch_results_{timestamp}.xlsx')
    excel_writer = ExcelResultWriter(TEMPLATE_PATH, output_xlsx)

    # 그리드 조합 생성
    param_names = list(GRID_PARAMS.keys())
    param_values = list(GRID_PARAMS.values())
    grid_combinations = list(product(*param_values))

    print(f"\nOutput: {exp_dir}")
    print(f"Grid combinations: {len(grid_combinations)}")
    for name, values in GRID_PARAMS.items():
        print(f"  {name}: {values}")
    print(f"{'='*80}")

    # 데이터 로드 (한 번만)
    print("\nLoading data...")
    train_data, train_labels, train_rr, train_pid, train_sid = load_or_extract_data(
        record_list=DS1_TRAIN_SPLIT, base_path=DATA_PATH, valid_leads=VALID_LEADS,
        out_len=OUT_LEN, split_name="Train_gridsearch"
    )
    valid_data, valid_labels, valid_rr, valid_pid, valid_sid = load_or_extract_data(
        record_list=DS1_VALID_SPLIT, base_path=DATA_PATH, valid_leads=VALID_LEADS,
        out_len=OUT_LEN, split_name="Valid_gridsearch"
    )
    test_data, test_labels, test_rr, test_pid, test_sid = load_or_extract_data(
        record_list=DS2_TEST, base_path=DATA_PATH, valid_leads=VALID_LEADS,
        out_len=OUT_LEN, split_name="Test_gridsearch"
    )

    # DataLoader 생성
    train_dataset = ECGDataset(train_data, train_rr, train_labels, train_pid, train_sid)
    valid_dataset = ECGDataset(valid_data, valid_rr, valid_labels, valid_pid, valid_sid)
    test_dataset = ECGDataset(test_data, test_rr, test_labels, test_pid, test_sid)

    train_loader = DataLoader(train_dataset, batch_size=BATCH_SIZE, shuffle=True,
                              num_workers=4, pin_memory=True)
    valid_loader = DataLoader(valid_dataset, batch_size=BATCH_SIZE, shuffle=False,
                              num_workers=4, pin_memory=True)
    test_loader = DataLoader(test_dataset, batch_size=BATCH_SIZE, shuffle=False,
                             num_workers=4, pin_memory=True)

    print(f"  Train: {len(train_labels):,} | Valid: {len(valid_labels):,} | Test: {len(test_labels):,}")

    data_loaders = (train_loader, valid_loader, test_loader)

    # 그리드 서치 실행
    results_list = []
    total_start = time.time()

    for trial_num, values in enumerate(grid_combinations, 1):
        params = dict(zip(param_names, values))

        result = run_single_experiment(
            params=params,
            trial_num=trial_num,
            total_trials=len(grid_combinations),
            device=device,
            data_loaders=data_loaders,
            excel_writer=excel_writer,
            exp_dir=exp_dir
        )
        results_list.append(result)

    # 최종 요약
    total_time = (time.time() - total_start) / 60
    successful = [r for r in results_list if r['status'] == 'success']

    print(f"\n{'='*80}")
    print(f"GRID SEARCH COMPLETE")
    print(f"{'='*80}")
    print(f"Total time: {total_time:.1f} min")
    print(f"Successful trials: {len(successful)}/{len(grid_combinations)}")

    if successful:
        # Best AUROC 찾기
        best_auroc = max(successful, key=lambda x: x['best_valid_auroc'])
        print(f"\nBest Valid AUROC: {best_auroc['best_valid_auroc']:.4f}")
        print(f"  Params: {best_auroc['params']}")

        # Best AUPRC 찾기
        best_auprc = max(successful, key=lambda x: x['best_valid_auprc'])
        print(f"\nBest Valid AUPRC: {best_auprc['best_valid_auprc']:.4f}")
        print(f"  Params: {best_auprc['params']}")

    print(f"\nResults saved to: {output_xlsx}")
    print(f"{'='*80}")

    return results_list, exp_dir


# =============================================================================
# 메인 실행
# =============================================================================

if __name__ == '__main__':
    results_list, exp_dir = run_gridsearch()
