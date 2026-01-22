"""
main_autoexp.py - Fusion Type 자동 실험 스크립트
MACNN_SE 클래스의 fusion_type (None, concat, concat_proj, mhca)에 대한 자동 실험

하이퍼파라미터 그리드 서치가 완료되었다고 가정하고,
기본 설정으로 fusion_type만 변경하여 실험 자동화

결과:
- model_fusion.xlsx 템플릿에 결과 저장 (Performance Metrics, Confusion 시트)
- Best AUROC, Best AUPRC 모델 각각 저장
"""

import os
import time
import copy
import shutil
from collections import Counter
from datetime import datetime

import numpy as np
import pandas as pd
import torch
from torch.utils.data import DataLoader
from openpyxl import load_workbook

from tqdm import tqdm

from utils import set_seed, load_or_extract_data
from model import MACNN_SE
from dataloader import ECGDataset
from train import train_one_epoch, validate, save_model
from evaluate_module import evaluate, calculate_metrics, save_confusion_matrix
from logger import (
    TrainingLogger, print_epoch_header, print_per_class_metrics,
    print_epoch_stats, print_confidence_stats, print_epoch_time
)

# =============================================================================
# 설정
# =============================================================================

DATA_PATH = './data/mit-bih-arrhythmia-database-1.0.0/'
OUTPUT_PATH = './fusion_results/'
TEMPLATE_PATH = './model_fusion.xlsx'
BATCH_SIZE = 1024
EPOCHS = 50
LR = 0.0001
WEIGHT_DECAY = 1e-3
SEED = 1234
POLY1_EPS = 0.0
POLY2_EPS = 0.0
CLASSES = ['N', 'S', 'V', 'F']

# RR Feature 설정
RR_FEATURE_OPTION = "opt3"
RR_FEATURE_DIMS = {"opt1": 7, "opt2": 38, "opt3": 7, "opt4": 7}

# ECG Parameters
VALID_LEADS = ['MLII', 'V1', 'V2', 'V4', 'V5']
OUT_LEN = 720

# 데이터 분할
DS1_TRAIN_SPLIT = [
    '101', '106', '108', '109', '112', '115', '116', '118', '119',
    '122', '201', '203', '209', '215', '223', '230', '208'
]
DS1_VALID_SPLIT = ['114', '124', '205', '207', '220']

DS2_TEST = [
    '100', '103', '105', '111', '113', '117', '121', '123', '200', '202',
    '210', '212', '213', '214', '219', '221', '222', '228', '231', '232',
    '233', '234'
]

# =============================================================================
# Fusion Type 실험 설정
# =============================================================================

# 실험할 fusion_type 리스트
FUSION_TYPES = [None, 'concat', 'concat_proj', 'mhca']

# num_heads 그리드 서치 값 (mhca에서만 사용)
NUM_HEADS_LIST = [1, 2, 4, 8]

# MACNN_SE 기본 설정 (그리드 서치 결과 또는 기본값 사용)
DEFAULT_MODEL_CONFIG = {
    'reduction': 16,
    'aspp_bn': True,
    'aspp_act': True,
    'lead': 1,
    'p': 0.0,
    'dilations': (1, 6, 12, 18),
    'act_func': 'tanh',
    'f_act_func': 'tanh',
    'apply_residual': False,
    'fusion_emb': 64,
    'fusion_expansion': 2,
    'rr_dim': RR_FEATURE_DIMS[RR_FEATURE_OPTION],
    'num_heads': 1,
}

# Fusion type별 실험 이름
FUSION_TYPE_NAMES = {
    None: 'Baseline',
    'concat': 'Concat',
    'concat_proj': 'ConcatProj',
    'mhca': 'MHCA',
}

# =============================================================================
# 엑셀 결과 저장 클래스
# =============================================================================

class ExcelResultWriter:
    """엑셀 템플릿에 결과를 실시간으로 저장하는 클래스"""

    def __init__(self, template_path, output_path):
        self.template_path = template_path
        self.output_path = output_path
        self.current_row = 3  # 데이터 시작 행 (0-indexed, 헤더 3줄 이후)
        self.confusion_start_row = 0  # Confusion 시트 시작 행

        # 템플릿 복사
        shutil.copy(template_path, output_path)
        print(f"Template copied to: {output_path}")

    def write_metrics(self, exp_name: str, metrics: dict, best_type: str):
        """
        Performance Metrics 시트에 결과 작성

        Args:
            exp_name: 실험 이름 (예: "MHCA_auroc")
            metrics: calculate_metrics 결과
            best_type: 'auroc' 또는 'auprc'
        """
        wb = load_workbook(self.output_path)
        ws = wb['Performance Metrics']

        row = self.current_row + 1  # openpyxl은 1-indexed

        # 실험명 (best_type 포함)
        full_name = f"{exp_name}_{best_type}"
        ws.cell(row=row, column=1, value=full_name)

        # Macro metrics (columns 2-6)
        ws.cell(row=row, column=2, value=metrics['macro_accuracy'])
        ws.cell(row=row, column=3, value=metrics['macro_recall'])  # Sensitivity
        ws.cell(row=row, column=4, value=metrics['macro_specificity'])
        ws.cell(row=row, column=5, value=metrics['macro_prec'])
        ws.cell(row=row, column=6, value=metrics['macro_f1'])

        # Weighted metrics (columns 7-11)
        ws.cell(row=row, column=7, value=metrics['weighted_accuracy'])
        ws.cell(row=row, column=8, value=metrics['weighted_recall'])
        ws.cell(row=row, column=9, value=metrics['weighted_specificity'])
        ws.cell(row=row, column=10, value=metrics['weighted_prec'])
        ws.cell(row=row, column=11, value=metrics['weighted_f1'])

        # Per-class metrics (N, S, V, F)
        for i, cls in enumerate(CLASSES):
            base_col = 12 + i * 5
            ws.cell(row=row, column=base_col, value=metrics['per_class_accuracy'][i])
            ws.cell(row=row, column=base_col + 1, value=metrics['per_class_recall'][i])
            ws.cell(row=row, column=base_col + 2, value=metrics['per_class_specificity'][i])
            ws.cell(row=row, column=base_col + 3, value=metrics['per_class_precision'][i])
            ws.cell(row=row, column=base_col + 4, value=metrics['per_class_f1'][i])

        wb.save(self.output_path)
        self.current_row += 1
        print(f"  Results written to row {row}: {full_name}")

    def write_confusion_matrix(self, exp_name: str, cm: np.ndarray, best_type: str):
        """
        Confusion 시트에 혼동 행렬 작성

        Args:
            exp_name: 실험 이름
            cm: 혼동 행렬 (4x4)
            best_type: 'auroc' 또는 'auprc'
        """
        wb = load_workbook(self.output_path)
        ws = wb['Confusion']

        # 새로운 confusion matrix 블록 시작 위치 찾기
        # 각 블록은 8행 (EXPNAME, 헤더2줄, 데이터4줄, 빈줄1개)
        block_start = self.confusion_start_row + 1  # openpyxl은 1-indexed

        full_name = f"{exp_name}_{best_type}"

        # EXPNAME 행
        ws.cell(row=block_start, column=1, value=full_name)

        # Predicted 헤더
        ws.cell(row=block_start + 1, column=3, value="Predicted")

        # 클래스 헤더
        for i, cls in enumerate(CLASSES):
            ws.cell(row=block_start + 2, column=3 + i, value=cls)

        # Actual 라벨 및 데이터
        ws.cell(row=block_start + 3, column=1, value="Actual")
        for i, cls in enumerate(CLASSES):
            ws.cell(row=block_start + 3 + i, column=2, value=cls)
            for j in range(4):
                ws.cell(row=block_start + 3 + i, column=3 + j, value=int(cm[i, j]))

        wb.save(self.output_path)
        self.confusion_start_row += 8  # 다음 블록 위치
        print(f"  Confusion matrix written for: {full_name}")


# =============================================================================
# 단일 실험 수행 함수
# =============================================================================

def run_single_experiment(fusion_type, exp_name: str, device: torch.device,
                          data_loaders: tuple, excel_writer: ExcelResultWriter,
                          exp_dir: str, num_heads: int = 1):
    """
    단일 fusion type 실험 수행

    Args:
        fusion_type: None, 'concat', 'concat_proj', 또는 'mhca'
        exp_name: 실험 이름
        device: torch device
        data_loaders: (train_loader, valid_loader, test_loader)
        excel_writer: 엑셀 결과 저장 객체
        exp_dir: 실험 디렉토리
        num_heads: attention head 수 (mhca에서 사용)

    Returns:
        result_dict: 결과 딕셔너리
    """
    train_loader, valid_loader, test_loader = data_loaders

    print(f"\n{'='*80}")
    print(f"Experiment: {exp_name} (fusion_type={fusion_type}, num_heads={num_heads})")
    print(f"{'='*80}")

    result_dict = {
        'exp_name': exp_name,
        'fusion_type': fusion_type,
        'num_heads': num_heads,
        'best_valid_auroc': 0.0,
        'best_valid_auprc': 0.0,
        'test_auroc_metrics': None,
        'test_auprc_metrics': None,
        'status': 'error'
    }

    try:
        set_seed(SEED)

        # MACNN_SE 모델 생성 - fusion_type과 num_heads 변경
        model = MACNN_SE(
            reduction=DEFAULT_MODEL_CONFIG['reduction'],
            aspp_bn=DEFAULT_MODEL_CONFIG['aspp_bn'],
            aspp_act=DEFAULT_MODEL_CONFIG['aspp_act'],
            lead=DEFAULT_MODEL_CONFIG['lead'],
            p=DEFAULT_MODEL_CONFIG['p'],
            dilations=DEFAULT_MODEL_CONFIG['dilations'],
            act_func=DEFAULT_MODEL_CONFIG['act_func'],
            f_act_func=DEFAULT_MODEL_CONFIG['f_act_func'],
            apply_residual=DEFAULT_MODEL_CONFIG['apply_residual'],
            fusion_type=fusion_type,
            fusion_emb=DEFAULT_MODEL_CONFIG['fusion_emb'],
            fusion_expansion=DEFAULT_MODEL_CONFIG['fusion_expansion'],
            rr_dim=DEFAULT_MODEL_CONFIG['rr_dim'],
            num_heads=num_heads
        ).to(device)

        # 파라미터 수 출력
        total_params = sum(p.numel() for p in model.parameters())
        print(f"  Model parameters: {total_params/1e6:.2f}M")

        # 학습 설정
        optimizer = torch.optim.AdamW(model.parameters(), lr=LR, weight_decay=WEIGHT_DECAY)
        scheduler = torch.optim.lr_scheduler.StepLR(optimizer, step_size=20, gamma=0.5)

        # Best 모델 추적
        best_auroc = {'value': 0.0, 'epoch': 0, 'state_dict': None}
        best_auprc = {'value': 0.0, 'epoch': 0, 'state_dict': None}

        # 학습 루프
        for epoch in range(1, EPOCHS + 1):
            current_lr = optimizer.param_groups[0]['lr']

            # Train
            train_loss, train_metrics, *p_t_train = train_one_epoch(
                model, train_loader, POLY1_EPS, POLY2_EPS, optimizer, device
            )

            print_epoch_stats(epoch, train_loss, train_metrics['acc'], current_lr, phase='Train')
            # print_per_class_metrics(train_metrics, CLASSES, phase='Train')
            # print_confidence_stats(*p_t_train, phase='Train')
            logger.log_epoch(epoch, train_loss, train_metrics, phase='train')
            logger.log_confidence(epoch, *p_t_train, phase='train')
            # Validation
            valid_loss, valid_metrics, *p_t_valid = validate(
                model, valid_loader, POLY1_EPS, POLY2_EPS, device
            )

            print_epoch_stats(epoch, valid_loss, valid_metrics['acc'], current_lr, phase='Valid')
            # print_per_class_metrics(valid_metrics, CLASSES, phase='Valid')
            # print_confidence_stats(*p_t_valid, phase='Valid')
            logger.log_epoch(epoch, valid_loss, valid_metrics, phase='valid')
            logger.log_confidence(epoch, *p_t_valid, phase='valid')

            # Best AUROC 체크
            if valid_metrics['macro_auroc'] > best_auroc['value']:
                best_auroc = {
                    'value': valid_metrics['macro_auroc'],
                    'epoch': epoch,
                    'state_dict': copy.deepcopy(model.state_dict())
                }
                print(f"  ★ [BEST AUROC] {best_auroc['value']:.4f}")

            # Best AUPRC 체크
            if valid_metrics['macro_auprc'] > best_auprc['value']:
                best_auprc = {
                    'value': valid_metrics['macro_auprc'],
                    'epoch': epoch,
                    'state_dict': copy.deepcopy(model.state_dict())
                }
                print(f"  ★ [BEST AUPRC] {best_auprc['value']:.4f}")

            scheduler.step()
            print("=" * 120 + "\n")

        # =====================================================================
        # Best AUROC 모델로 테스트
        # =====================================================================
        print(f"\n  Testing Best AUROC model (epoch {best_auroc['epoch']})...")
        model.load_state_dict(best_auroc['state_dict'])
        model.eval()

        y_pred_auroc, y_true_auroc, _ = evaluate(model, test_loader, device)
        metrics_auroc = calculate_metrics(np.array(y_true_auroc), np.array(y_pred_auroc))

        # 모델 저장
        auroc_path = os.path.join(exp_dir, 'best_weights', f'{exp_name}_best_auroc.pth')
        torch.save({
            'model_state_dict': best_auroc['state_dict'],
            'epoch': best_auroc['epoch'],
            'fusion_type': fusion_type,
            'num_heads': num_heads,
            'auroc': best_auroc['value']
        }, auroc_path)

        # Confusion matrix 이미지 저장
        cm_auroc_path = os.path.join(exp_dir, 'confusion_matrices', f'{exp_name}_auroc.png')
        save_confusion_matrix(metrics_auroc['confusion_matrix'], CLASSES, cm_auroc_path)

        # 엑셀에 결과 작성
        excel_writer.write_metrics(exp_name, metrics_auroc, 'auroc')
        excel_writer.write_confusion_matrix(exp_name, metrics_auroc['confusion_matrix'], 'auroc')

        print(f"  AUROC Test - Macro F1: {metrics_auroc['macro_f1']:.4f}, "
              f"Acc: {metrics_auroc['overall_accuracy']:.4f}")

        # =====================================================================
        # Best AUPRC 모델로 테스트
        # =====================================================================
        print(f"\n  Testing Best AUPRC model (epoch {best_auprc['epoch']})...")
        model.load_state_dict(best_auprc['state_dict'])
        model.eval()

        y_pred_auprc, y_true_auprc, _ = evaluate(model, test_loader, device)
        metrics_auprc = calculate_metrics(np.array(y_true_auprc), np.array(y_pred_auprc))

        # 모델 저장
        auprc_path = os.path.join(exp_dir, 'best_weights', f'{exp_name}_best_auprc.pth')
        torch.save({
            'model_state_dict': best_auprc['state_dict'],
            'epoch': best_auprc['epoch'],
            'fusion_type': fusion_type,
            'num_heads': num_heads,
            'auprc': best_auprc['value']
        }, auprc_path)

        # Confusion matrix 이미지 저장
        cm_auprc_path = os.path.join(exp_dir, 'confusion_matrices', f'{exp_name}_auprc.png')
        save_confusion_matrix(metrics_auprc['confusion_matrix'], CLASSES, cm_auprc_path)

        # 엑셀에 결과 작성
        excel_writer.write_metrics(exp_name, metrics_auprc, 'auprc')
        excel_writer.write_confusion_matrix(exp_name, metrics_auprc['confusion_matrix'], 'auprc')

        print(f"  AUPRC Test - Macro F1: {metrics_auprc['macro_f1']:.4f}, "
              f"Acc: {metrics_auprc['overall_accuracy']:.4f}")

        # 결과 저장
        result_dict['best_valid_auroc'] = best_auroc['value']
        result_dict['best_valid_auprc'] = best_auprc['value']
        result_dict['test_auroc_metrics'] = metrics_auroc
        result_dict['test_auprc_metrics'] = metrics_auprc
        result_dict['status'] = 'success'

    except Exception as e:
        print(f"  Error: {str(e)}")
        import traceback
        traceback.print_exc()
        result_dict['status'] = 'error'
   
    return result_dict


# =============================================================================
# 메인 자동 실험 함수
# =============================================================================

def run_auto_experiments():
    """전체 fusion type 자동 실험 실행"""
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

    print(f"\n{'='*80}")
    print(f"MACNN_SE Fusion Type Auto Experiment")
    print(f"{'='*80}")
    print(f"Device: {device}")
    if torch.cuda.is_available():
        print(f"GPU: {torch.cuda.get_device_name(0)}")

    # 실험 디렉토리 생성
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    exp_dir = os.path.join(OUTPUT_PATH, f'fusion_exp_{timestamp}')
    os.makedirs(exp_dir, exist_ok=True)
    os.makedirs(os.path.join(exp_dir, 'best_weights'), exist_ok=True)
    os.makedirs(os.path.join(exp_dir, 'confusion_matrices'), exist_ok=True)

    # 엑셀 결과 저장 객체 생성
    output_xlsx = os.path.join(exp_dir, f'fusion_results_{timestamp}.xlsx')
    excel_writer = ExcelResultWriter(TEMPLATE_PATH, output_xlsx)

    print(f"\nOutput: {exp_dir}")
    print(f"Fusion types to test: {FUSION_TYPES}")
    print(f"num_heads grid search (for mhca): {NUM_HEADS_LIST}")
    print(f"Default config: {DEFAULT_MODEL_CONFIG}")
    print(f"{'='*80}")

    # 데이터 로드 (한 번만)
    print("\nLoading data...")
    train_data, train_labels, train_rr, train_pid, train_sid = load_or_extract_data(
        record_list=DS1_TRAIN_SPLIT, base_path=DATA_PATH, valid_leads=VALID_LEADS,
        out_len=OUT_LEN, split_name="Train_fusion"
    )
    valid_data, valid_labels, valid_rr, valid_pid, valid_sid = load_or_extract_data(
        record_list=DS1_VALID_SPLIT, base_path=DATA_PATH, valid_leads=VALID_LEADS,
        out_len=OUT_LEN, split_name="Valid_fusion"
    )
    test_data, test_labels, test_rr, test_pid, test_sid = load_or_extract_data(
        record_list=DS2_TEST, base_path=DATA_PATH, valid_leads=VALID_LEADS,
        out_len=OUT_LEN, split_name="Test_fusion"
    )

    # DataLoader 생성
    train_dataset = ECGDataset(train_data, train_rr, train_labels, train_pid, train_sid)
    valid_dataset = ECGDataset(valid_data, valid_rr, valid_labels, valid_pid, valid_sid)
    test_dataset = ECGDataset(test_data, test_rr, test_labels, test_pid, test_sid)

    train_loader = DataLoader(train_dataset, batch_size=BATCH_SIZE, shuffle=True,
                              num_workers=4, pin_memory=True)
    valid_loader = DataLoader(valid_dataset, batch_size=BATCH_SIZE, shuffle=False,
                              num_workers=4, pin_memory=True)
    test_loader = DataLoader(test_dataset, batch_size=BATCH_SIZE, shuffle=False,
                             num_workers=4, pin_memory=True)

    print(f"  Train: {len(train_labels):,} samples | {dict(Counter(train_labels))}")
    print(f"  Valid: {len(valid_labels):,} samples | {dict(Counter(valid_labels))}")
    print(f"  Test : {len(test_labels):,} samples | {dict(Counter(test_labels))}")

    data_loaders = (train_loader, valid_loader, test_loader)

    # 자동 실험 실행
    results_list = []
    total_start = time.time()

    for fusion_type in FUSION_TYPES:
        base_exp_name = FUSION_TYPE_NAMES[fusion_type]

        # mhca의 경우 num_heads 그리드 서치 수행
        if fusion_type == 'mhca':
            for num_heads in NUM_HEADS_LIST:
                exp_name = f"{base_exp_name}_h{num_heads}"
                result = run_single_experiment(
                    fusion_type=fusion_type,
                    exp_name=exp_name,
                    device=device,
                    data_loaders=data_loaders,
                    excel_writer=excel_writer,
                    exp_dir=exp_dir,
                    num_heads=num_heads
                )
                results_list.append(result)
        else:
            # 다른 fusion type은 기본 num_heads=1 사용
            result = run_single_experiment(
                fusion_type=fusion_type,
                exp_name=base_exp_name,
                device=device,
                data_loaders=data_loaders,
                excel_writer=excel_writer,
                exp_dir=exp_dir,
                num_heads=1
            )
            results_list.append(result)

    # 최종 요약
    total_time = (time.time() - total_start) / 60
    successful = [r for r in results_list if r['status'] == 'success']

    print(f"\n{'='*80}")
    print(f"AUTO EXPERIMENT COMPLETE")
    print(f"{'='*80}")
    print(f"Total time: {total_time:.1f} min")
    print(f"Successful experiments: {len(successful)}/{len(FUSION_TYPES)}")

    # 총 실험 수 계산
    total_experiments = len(FUSION_TYPES) - 1 + len(NUM_HEADS_LIST)  # mhca 제외한 fusion types + mhca * num_heads

    print(f"\nResults Summary:")
    print(f"{'Experiment':<20} {'num_heads':<10} {'Valid AUROC':<12} {'Valid AUPRC':<12} {'Test F1 (AUROC)':<15} {'Test F1 (AUPRC)':<15}")
    print("-" * 85)

    for result in results_list:
        if result['status'] == 'success':
            auroc_f1 = result['test_auroc_metrics']['macro_f1']
            auprc_f1 = result['test_auprc_metrics']['macro_f1']
            print(f"{result['exp_name']:<20} {result['num_heads']:<10} {result['best_valid_auroc']:<12.4f} "
                  f"{result['best_valid_auprc']:<12.4f} {auroc_f1:<15.4f} {auprc_f1:<15.4f}")
        else:
            print(f"{result['exp_name']:<20} FAILED")

    print(f"\nResults saved to: {output_xlsx}")
    print(f"{'='*80}")

    return results_list, exp_dir


# =============================================================================
# 메인 실행
# =============================================================================

if __name__ == '__main__':
    results_list, exp_dir = run_auto_experiments()
