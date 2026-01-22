# =============================================================================
# main.py - ECG Classification Training (MACNN_SE + DAEAC / Standard)
# =============================================================================

import os
import time
import numpy as np
import torch
from torch.utils.data import DataLoader

import config
from utils import set_seed, load_or_extract_data
from model import MACNN_SE
from dataloader import ECGDataset, DAEACDataset
from train import train_one_epoch, validate, save_model
from evaluate_module import evaluate, calculate_metrics, print_metrics
from logger import TrainingLogger, print_epoch_header, print_epoch_stats

# ==========================================s===================================
# 1. Initialization
# =============================================================================
set_seed(config.SEED)
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

exp_dir = config.create_experiment_dir()
model_weights_dir = os.path.join(exp_dir, "model_weights")
best_weights_dir = os.path.join(exp_dir, "best_weights")
os.makedirs(model_weights_dir, exist_ok=True)
os.makedirs(best_weights_dir, exist_ok=True)

logger = TrainingLogger(os.path.join(exp_dir, "runs"))

print("\n" + "=" * 80)
print(f"MODE   : {'DAEAC (Paper Style)' if config.is_daeac_style() else 'Standard'}")
print(f"STYLE  : {config.CURR_STYLE}")
print(f"LEAD   : {config.CURR_LEAD}")
print(f"LENGTH : {config.CURR_OUT_LEN}")
print(f"DEVICE : {device}")
print("=" * 80)

# =============================================================================
# 2. Data Loading
# =============================================================================
print("\n[1/3] Loading Data...")

def get_loader(records, split_name, shuffle=False):
    data = load_or_extract_data(
        record_list=records,
        base_path=config.DATA_PATH,
        valid_leads=config.VALID_LEADS,
        out_len=config.CURR_OUT_LEN,
        split_name=split_name,
        extraction_style=config.CURR_STYLE
    )
    DS = DAEACDataset if config.is_daeac_style() else ECGDataset
    return DataLoader(
        DS(*data),
        batch_size=config.BATCH_SIZE,
        shuffle=shuffle,
        num_workers=4,
        pin_memory=True
    )

train_loader = get_loader(config.DS1_TRAIN, "Train", shuffle=True)
valid_loader = get_loader(config.DS1_VALID, "Valid")
test_loader  = get_loader(config.DS2_TEST,  "Test")

# =============================================================================
# 3. Model
# =============================================================================
print("\n[2/3] Building Model...")

model = MACNN_SE(
    rr_dim=config.RR_FEATURE_DIMS[config.RR_FEATURE_OPTION],
    **config.MACNN_SE_CONFIG
).to(device)

optimizer = torch.optim.Adam(
    model.parameters(),
    lr=config.LR,
    weight_decay=config.WEIGHT_DECAY
)

scheduler = torch.optim.lr_scheduler.LambdaLR(
    optimizer,
    lr_lambda=lambda step: 0.99 ** (step / 200)
)

# =============================================================================
# 4. Class Weights
# =============================================================================
labels = train_loader.dataset.labels
labels = labels.cpu().numpy() if torch.is_tensor(labels) else np.array(labels)

cnt = np.bincount(labels.flatten().astype(np.int64), minlength=len(config.CLASSES))
weights = torch.tensor(1.0 / (cnt + 1e-8), dtype=torch.float32).to(device)
weights = weights / weights.sum() * len(config.CLASSES)

print(f"Class Count  : {cnt}")
print(f"Class Weight : {weights.cpu().numpy()}")

# =============================================================================
# 5. Best Model Tracker
# =============================================================================
best = {
    'auprc':  {'value': 0.0, 'epoch': 0, 'path': None},
    'auroc':  {'value': 0.0, 'epoch': 0, 'path': None},
    'recall': {'value': 0.0, 'epoch': 0, 'path': None},
    'last':   {'value': 0.0, 'epoch': 0, 'path': None},
}

# =============================================================================
# 6. Training Loop
# =============================================================================
print("\n[3/3] Training Start")
print_epoch_header()

total_start = time.time()

for epoch in range(1, config.EPOCHS + 1):
    epoch_start = time.time()

    # ---- Train ----
    tr_loss, tr_met, *_ = train_one_epoch(
        model, train_loader, epoch,
        optimizer, device, scheduler, weights
    )

    # ---- Valid ----
    va_loss, va_met, *_ = validate(
        model, valid_loader, device
    )

    lr = optimizer.param_groups[0]['lr']
    print_epoch_stats(epoch, tr_loss, tr_met['acc'], lr, 'Train')
    print_epoch_stats(epoch, va_loss, va_met['acc'], lr, 'Valid')

    logger.log_epoch(epoch, tr_loss, tr_met, 'train')
    logger.log_epoch(epoch, va_loss, va_met, 'valid')

    # ---- Save all epochs ----
    save_model(
        model, optimizer, epoch, va_met,
        os.path.join(model_weights_dir, f"{config.EXP_NAME}_epoch_{epoch}.pth")
    )

    # ---- Best selection ----
    if va_met['macro_auprc'] > best['auprc']['value']:
        best['auprc'] = {
            'value': va_met['macro_auprc'],
            'epoch': epoch,
            'path': os.path.join(best_weights_dir, f"best_auprc_{config.EXP_NAME}.pth")
        }
        save_model(model, optimizer, epoch, va_met, best['auprc']['path'])

    if va_met['macro_auroc'] > best['auroc']['value']:
        best['auroc'] = {
            'value': va_met['macro_auroc'],
            'epoch': epoch,
            'path': os.path.join(best_weights_dir, f"best_auroc_{config.EXP_NAME}.pth")
        }
        save_model(model, optimizer, epoch, va_met, best['auroc']['path'])

    if va_met['macro_recall'] > best['recall']['value']:
        best['recall'] = {
            'value': va_met['macro_recall'],
            'epoch': epoch,
            'path': os.path.join(best_weights_dir, f"best_recall_{config.EXP_NAME}.pth")
        }
        save_model(model, optimizer, epoch, va_met, best['recall']['path'])

    if epoch == config.EPOCHS:
        best['last'] = {
            'value': va_met['acc'],
            'epoch': epoch,
            'path': os.path.join(best_weights_dir, f"best_last_{config.EXP_NAME}.pth")
        }
        save_model(model, optimizer, epoch, va_met, best['last']['path'])

    print(f"Epoch Time: {time.time() - epoch_start:.1f}s")
    print("-" * 120)

logger.close()

# =============================================================================
# 7. Test Evaluation (Best Models)
# =============================================================================
print("\n" + "=" * 80)
print("Test Set Evaluation")
print("=" * 80)

for tag, info in best.items():
    if info['path'] is None or not os.path.exists(info['path']):
        continue

    print(f"\n--- BEST {tag.upper()} | Epoch {info['epoch']} ---")

    ckpt = torch.load(info['path'], map_location=device)
    model.load_state_dict(ckpt['model_state_dict'])
    model.eval()

    y_pred, y_true, _ = evaluate(model, test_loader, device)
    metrics = calculate_metrics(y_true, y_pred)
    print_metrics(metrics, config.CLASSES)

print("\n" + "=" * 80)
print(f"ALL RESULTS SAVED TO: {exp_dir}")
print("=" * 80)


# =============================================================================
# 엑셀 결과 저장 클래스
# =============================================================================

import shutil
import numpy as np
from openpyxl import load_workbook

class ExcelResultWriter:
    """
    엑셀 템플릿에 실험 결과를 누적 저장하는 클래스
    - Performance Metrics 시트
    - Confusion 시트
    """

    def __init__(self, template_path, output_path):
        self.template_path = template_path
        self.output_path = output_path
        self.current_row = 3          # Performance Metrics 데이터 시작 행 (0-index 기준)
        self.confusion_start_row = 0  # Confusion 시트 블록 시작 행

        # 템플릿 복사
        shutil.copy(template_path, output_path)
        print(f"[Excel] Template copied to: {output_path}")

    def write_metrics(self, exp_name: str, metrics: dict, best_type: str):
        """
        Performance Metrics 시트에 결과 기록

        Args:
            exp_name  : 실험 이름
            metrics   : calculate_metrics 결과 dict
            best_type : 'auroc' | 'auprc' | 'recall' | 'last'
        """
        wb = load_workbook(self.output_path)
        ws = wb['Performance Metrics']

        row = self.current_row + 1  # openpyxl은 1-index

        full_name = f"{exp_name}_{best_type}"
        ws.cell(row=row, column=1, value=full_name)

        # ---- Macro ----
        ws.cell(row=row, column=2, value=metrics['macro_accuracy'])
        ws.cell(row=row, column=3, value=metrics['macro_recall'])
        ws.cell(row=row, column=4, value=metrics['macro_specificity'])
        ws.cell(row=row, column=5, value=metrics['macro_prec'])
        ws.cell(row=row, column=6, value=metrics['macro_f1'])

        # ---- Weighted ----
        ws.cell(row=row, column=7,  value=metrics['weighted_accuracy'])
        ws.cell(row=row, column=8,  value=metrics['weighted_recall'])
        ws.cell(row=row, column=9,  value=metrics['weighted_specificity'])
        ws.cell(row=row, column=10, value=metrics['weighted_prec'])
        ws.cell(row=row, column=11, value=metrics['weighted_f1'])

        # ---- Per-class ----
        for i, cls in enumerate(CLASSES):
            base_col = 12 + i * 5
            ws.cell(row=row, column=base_col,     value=metrics['per_class_accuracy'][i])
            ws.cell(row=row, column=base_col + 1, value=metrics['per_class_recall'][i])
            ws.cell(row=row, column=base_col + 2, value=metrics['per_class_specificity'][i])
            ws.cell(row=row, column=base_col + 3, value=metrics['per_class_precision'][i])
            ws.cell(row=row, column=base_col + 4, value=metrics['per_class_f1'][i])

        wb.save(self.output_path)
        self.current_row += 1

        print(f"[Excel] Metrics written → row {row}: {full_name}")

    def write_confusion_matrix(self, exp_name: str, cm: np.ndarray, best_type: str):
        """
        Confusion 시트에 confusion matrix 블록 단위로 기록

        Args:
            exp_name  : 실험 이름
            cm        : (C x C) confusion matrix
            best_type : 'auroc' | 'auprc' | 'recall' | 'last'
        """
        wb = load_workbook(self.output_path)
        ws = wb['Confusion']

        block_start = self.confusion_start_row + 1
        full_name = f"{exp_name}_{best_type}"

        # 제목
        ws.cell(row=block_start, column=1, value=full_name)

        # Predicted header
        ws.cell(row=block_start + 1, column=3, value="Predicted")

        # Class header
        for i, cls in enumerate(CLASSES):
            ws.cell(row=block_start + 2, column=3 + i, value=cls)

        # Actual + matrix
        ws.cell(row=block_start + 3, column=1, value="Actual")
        for i, cls in enumerate(CLASSES):
            ws.cell(row=block_start + 3 + i, column=2, value=cls)
            for j in range(len(CLASSES)):
                ws.cell(
                    row=block_start + 3 + i,
                    column=3 + j,
                    value=int(cm[i, j])
                )

        wb.save(self.output_path)
        self.confusion_start_row += 8

        print(f"[Excel] Confusion matrix written: {full_name}")


excel_writer = ExcelResultWriter(
    template_path="model_gridsearch.xlsx",
    output_path=os.path.join(exp_dir, f"results_{config.EXP_NAME}.xlsx")
)

excel_writer.write_metrics(config.EXP_NAME, metrics, tag)
excel_writer.write_confusion_matrix(
    config.EXP_NAME,
    metrics['confusion_matrix'],
    tag
)