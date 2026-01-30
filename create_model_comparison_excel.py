"""
모델 비교 결과를 Excel로 저장
"""

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

CLASSES = ["N", "S", "V", "F"]

# 모델별 데이터 (Per-class: Acc, Sens, Spec, Prec, F1)
MODEL_DATA = {
    "A0 (ResUDense)": {
        "per_class": {
            "N": [93.36, 98.51, 51.51, 94.28, 96.35],
            "S": [95.62, 3.05, 99.17, 12.39, 4.89],
            "V": [98.25, 79.00, 99.58, 92.88, 85.38],
            "F": [98.68, 0.77, 99.45, 1.10, 0.91],
        },
        "macro": [96.48, 45.33, 87.43, 50.16, 46.88],
        "weighted": [93.80, 92.95, 56.76, 90.44, 91.51],
    },
    "A1 (Naive Concat)": {
        "per_class": {
            "N": [91.50, 95.32, 60.51, 95.15, 95.23],
            "S": [95.24, 8.33, 98.58, 18.35, 11.46],
            "V": [96.52, 91.21, 96.89, 67.02, 77.26],
            "F": [98.94, 1.55, 99.71, 4.00, 2.23],
        },
        "macro": [95.55, 49.10, 88.92, 46.13, 46.55],
        "weighted": [92.02, 91.10, 64.58, 89.77, 90.24],
    },
    "A2 (Cross-Attention)": {
        "per_class": {
            "N": [93.89, 97.43, 65.12, 95.78, 96.60],
            "S": [96.74, 25.15, 99.49, 65.53, 36.35],
            "V": [96.32, 80.06, 97.45, 68.52, 73.84],
            "F": [98.82, 2.06, 99.58, 3.76, 2.66],
        },
        "macro": [96.44, 51.18, 90.41, 58.40, 52.36],
        "weighted": [94.19, 92.89, 68.75, 92.17, 92.16],
    },
    "B0 (ResUFormer)": {
        "per_class": {
            "N": [92.98, 96.60, 63.54, 95.56, 96.08],
            "S": [94.83, 14.26, 97.93, 20.93, 16.96],
            "V": [98.00, 90.84, 98.50, 80.73, 85.48],
            "F": [99.05, 0.52, 99.83, 2.33, 0.84],
        },
        "macro": [96.22, 50.55, 89.95, 49.88, 49.84],
        "weighted": [93.42, 92.43, 67.36, 91.11, 91.72],
    },
    "B1 (Naive Concat TF)": {
        "per_class": {
            "N": [96.04, 98.17, 78.75, 97.40, 97.79],
            "S": [96.71, 37.13, 99.00, 58.84, 45.53],
            "V": [98.57, 94.53, 98.85, 85.02, 89.53],
            "F": [98.64, 9.02, 99.35, 9.83, 9.41],
        },
        "macro": [97.49, 59.71, 93.99, 62.78, 60.56],
        "weighted": [96.25, 94.98, 80.96, 94.49, 94.63],
    },
    "B2 (Cross-Attention TF)": {
        "per_class": {
            "N": [94.75, 96.02, 84.42, 98.04, 97.02],
            "S": [97.17, 73.54, 98.08, 59.52, 65.79],
            "V": [98.33, 88.29, 99.02, 86.25, 87.26],
            "F": [98.18, 35.05, 98.67, 17.22, 23.09],
        },
        "macro": [97.11, 73.23, 95.05, 65.26, 68.29],
        "weighted": [95.10, 94.21, 85.99, 95.22, 94.65],
    },
    "C0 (MACNN-SE)": {
        "per_class": {
            "N": [93.56, 94.71, 84.21, 97.99, 96.32],
            "S": [95.96, 34.08, 98.33, 43.99, 38.40],
            "V": [96.24, 81.90, 97.24, 67.26, 73.86],
            "F": [96.69, 42.01, 97.12, 10.31, 16.56],
        },
        "macro": [95.61, 63.17, 94.23, 54.89, 56.29],
        "weighted": [93.85, 91.23, 85.68, 93.32, 92.10],
    },
    "C1 (Naive Concat)": {
        "per_class": {
            "N": [94.54, 99.17, 56.94, 94.92, 97.00],
            "S": [96.32, 0.98, 99.99, 75.00, 1.93],
            "V": [96.81, 76.17, 98.24, 74.96, 75.56],
            "F": [98.91, 2.58, 99.67, 5.75, 3.56],
        },
        "macro": [96.64, 44.72, 88.71, 62.66, 44.51],
        "weighted": [94.78, 93.29, 61.55, 92.20, 91.36],
    },
    "C0' (Early Ablation)": {
        "per_class": {
            "N": [92.54, 96.49, 60.47, 95.20, 95.84],
            "S": [90.48, 24.66, 93.00, 11.93, 16.08],
            "V": [93.21, 0.28, 99.66, 5.36, 0.53],
            "F": [97.46, 0.52, 98.23, 0.23, 0.32],
        },
        "macro": [93.42, 30.49, 87.84, 28.18, 28.19],
        "weighted": [92.55, 86.85, 64.51, 85.55, 85.96],
    },
    "C2' (Early Ablation)": {
        "per_class": {
            "N": [92.05, 95.92, 60.62, 95.19, 95.55],
            "S": [96.33, 2.45, 99.93, 58.44, 4.70],
            "V": [97.10, 84.44, 97.98, 74.36, 79.08],
            "F": [96.56, 6.44, 97.27, 1.82, 2.84],
        },
        "macro": [95.51, 47.31, 88.95, 57.45, 45.54],
        "weighted": [92.57, 91.02, 64.78, 91.75, 90.40],
    },
    "C2 (Cross-Attention)": {
        "per_class": {
            "N": [94.36, 96.88, 73.90, 96.79, 96.83],
            "S": [97.78, 55.20, 99.42, 78.48, 64.81],
            "V": [97.67, 79.19, 98.95, 83.96, 81.50],
            "F": [97.32, 17.27, 97.95, 6.23, 9.15],
        },
        "macro": [96.78, 62.13, 92.55, 66.36, 63.08],
        "weighted": [94.72, 93.57, 76.65, 94.57, 93.97],
    },
}


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Performance"

    # 스타일 설정
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Header 설정
    ws.cell(1, 1).value = "Model"
    header = ["Acc", "Sens", "Spec", "Prec", "F1"]
    col = 2
    for group in ["Macro", "Weighted"] + CLASSES:
        for h in header:
            ws.cell(1, col).value = f"{group}_{h}"
            ws.cell(1, col).font = header_font
            ws.cell(1, col).alignment = center_align
            ws.cell(1, col).fill = header_fill
            ws.cell(1, col).border = thin_border
            col += 1

    ws.cell(1, 1).font = header_font
    ws.cell(1, 1).fill = header_fill
    ws.cell(1, 1).border = thin_border

    # 데이터 입력
    for row_idx, (model_name, data) in enumerate(MODEL_DATA.items(), 2):
        ws.cell(row_idx, 1).value = model_name
        ws.cell(row_idx, 1).border = thin_border

        col = 2
        # Macro
        for val in data["macro"]:
            ws.cell(row_idx, col).value = f"{val:.2f}%"
            ws.cell(row_idx, col).alignment = center_align
            ws.cell(row_idx, col).border = thin_border
            col += 1

        # Weighted
        for val in data["weighted"]:
            ws.cell(row_idx, col).value = f"{val:.2f}%"
            ws.cell(row_idx, col).alignment = center_align
            ws.cell(row_idx, col).border = thin_border
            col += 1

        # Per-class
        for cls in CLASSES:
            for val in data["per_class"][cls]:
                ws.cell(row_idx, col).value = f"{val:.2f}%"
                ws.cell(row_idx, col).alignment = center_align
                ws.cell(row_idx, col).border = thin_border
                col += 1

    # 열 너비 조정
    from openpyxl.utils import get_column_letter
    ws.column_dimensions['A'].width = 25
    for col_idx in range(2, col):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    save_path = "/home/work/Ryuha/Model_Comparison_Results.xlsx"
    wb.save(save_path)
    print(f"Excel saved to: {save_path}")


if __name__ == "__main__":
    main()
