"""
Excel Result Writer

실험 결과를 엑셀 템플릿에 저장하는 유틸리티 클래스입니다.

사용 예시:
    from src.utils import ExcelResultWriter

    writer = ExcelResultWriter("template.xlsx", "output.xlsx")
    writer.write_metrics("exp_name", metrics, "auprc")
    writer.write_confusion_matrix("exp_name", cm, "auprc")

누적 기록 예시:
    from src.utils import CumulativeExcelWriter

    writer = CumulativeExcelWriter("template.xlsx", "cumulative_results.xlsx")
    writer.append_result("exp_name", metrics, config_dict)  # 자동으로 다음 row에 추가
"""

import os
import shutil
import numpy as np
from openpyxl import load_workbook
from datetime import datetime


class ExcelResultWriter:
    """
    엑셀 템플릿에 실험 결과를 저장하는 클래스

    Args:
        template_path: 엑셀 템플릿 파일 경로
        output_path: 결과 저장 경로
        classes: 클래스 이름 리스트 (default: ['N', 'S', 'V', 'F'])
    """

    def __init__(self, template_path: str, output_path: str, classes=None):
        self.template_path = template_path
        self.output_path = output_path
        self.classes = classes or ['N', 'S', 'V', 'F']
        self.current_row = 3  # 데이터 시작 행 (0-indexed, 헤더 3줄 이후)
        self.confusion_start_row = 0  # Confusion 시트 시작 행

        # 템플릿 복사
        shutil.copy(template_path, output_path)
        print(f"[Excel] Template copied to: {output_path}")

    def write_metrics(self, exp_name: str, metrics: dict, best_type: str):
        """
        Performance Metrics 시트에 결과 작성

        Args:
            exp_name: 실험 이름
            metrics: calculate_metrics 결과 dict
            best_type: 'auroc', 'auprc', 'recall' 등
        """
        wb = load_workbook(self.output_path)
        ws = wb['Performance Metrics']

        row = self.current_row + 1  # openpyxl은 1-indexed

        # 실험명 (best_type 포함)
        full_name = f"{exp_name}_{best_type}"
        ws.cell(row=row, column=1, value=full_name)

        # Macro metrics (columns 2-6)
        ws.cell(row=row, column=2, value=metrics.get('macro_accuracy', metrics.get('acc', 0)))
        ws.cell(row=row, column=3, value=metrics.get('macro_recall', 0))
        ws.cell(row=row, column=4, value=metrics.get('macro_specificity', 0))
        ws.cell(row=row, column=5, value=metrics.get('macro_precision', metrics.get('macro_prec', 0)))
        ws.cell(row=row, column=6, value=metrics.get('macro_f1', 0))

        # Weighted metrics (columns 7-11)
        ws.cell(row=row, column=7, value=metrics.get('weighted_accuracy', metrics.get('acc', 0)))
        ws.cell(row=row, column=8, value=metrics.get('weighted_recall', 0))
        ws.cell(row=row, column=9, value=metrics.get('weighted_specificity', 0))
        ws.cell(row=row, column=10, value=metrics.get('weighted_precision', metrics.get('weighted_prec', 0)))
        ws.cell(row=row, column=11, value=metrics.get('weighted_f1', 0))

        # Per-class metrics (N, S, V, F)
        per_class_acc = metrics.get('per_class_acc', metrics.get('per_class_accuracy', [0]*4))
        per_class_recall = metrics.get('per_class_recall', [0]*4)
        per_class_spec = metrics.get('per_class_specificity', [0]*4)
        per_class_prec = metrics.get('per_class_precision', [0]*4)
        per_class_f1 = metrics.get('per_class_f1', [0]*4)

        for i, cls in enumerate(self.classes):
            base_col = 12 + i * 5
            ws.cell(row=row, column=base_col, value=per_class_acc[i] if i < len(per_class_acc) else 0)
            ws.cell(row=row, column=base_col + 1, value=per_class_recall[i] if i < len(per_class_recall) else 0)
            ws.cell(row=row, column=base_col + 2, value=per_class_spec[i] if i < len(per_class_spec) else 0)
            ws.cell(row=row, column=base_col + 3, value=per_class_prec[i] if i < len(per_class_prec) else 0)
            ws.cell(row=row, column=base_col + 4, value=per_class_f1[i] if i < len(per_class_f1) else 0)

        wb.save(self.output_path)
        self.current_row += 1
        print(f"[Excel] Metrics written → row {row}: {full_name}")

    def write_confusion_matrix(self, exp_name: str, cm: np.ndarray, best_type: str):
        """
        Confusion 시트에 혼동 행렬 작성

        Args:
            exp_name: 실험 이름
            cm: 혼동 행렬 (4x4)
            best_type: 'auroc', 'auprc', 'recall' 등
        """
        wb = load_workbook(self.output_path)
        ws = wb['Confusion']

        # 각 블록은 8행 (EXPNAME, 헤더2줄, 데이터4줄, 빈줄1개)
        block_start = self.confusion_start_row + 1  # openpyxl은 1-indexed

        full_name = f"{exp_name}_{best_type}"

        # EXPNAME 행
        ws.cell(row=block_start, column=1, value=full_name)

        # Predicted 헤더
        ws.cell(row=block_start + 1, column=3, value="Predicted")

        # 클래스 헤더
        for i, cls in enumerate(self.classes):
            ws.cell(row=block_start + 2, column=3 + i, value=cls)

        # Actual 라벨 및 데이터
        ws.cell(row=block_start + 3, column=1, value="Actual")
        for i, cls in enumerate(self.classes):
            ws.cell(row=block_start + 3 + i, column=2, value=cls)
            for j in range(len(self.classes)):
                ws.cell(row=block_start + 3 + i, column=3 + j, value=int(cm[i, j]))

        wb.save(self.output_path)
        self.confusion_start_row += 8  # 다음 블록 위치
        print(f"[Excel] Confusion matrix written: {full_name}")

    def write_summary_row(self, exp_name: str, summary: dict):
        """
        간단한 요약 행 작성 (커스텀 컬럼)

        Args:
            exp_name: 실험 이름
            summary: {column_name: value} 형태의 dict
        """
        wb = load_workbook(self.output_path)
        ws = wb['Performance Metrics']

        row = self.current_row + 1
        ws.cell(row=row, column=1, value=exp_name)

        col = 2
        for key, value in summary.items():
            ws.cell(row=row, column=col, value=value)
            col += 1

        wb.save(self.output_path)
        self.current_row += 1
        print(f"[Excel] Summary written → row {row}: {exp_name}")


class CumulativeExcelWriter:
    """
    누적 엑셀 기록 클래스

    모든 실험 결과를 하나의 엑셀 파일에 계속 추가합니다.
    파일이 없으면 템플릿에서 생성하고, 있으면 마지막 row 다음부터 기록합니다.

    Args:
        template_path: 엑셀 템플릿 파일 경로
        output_path: 누적 결과 저장 경로
        classes: 클래스 이름 리스트 (default: ['N', 'S', 'V', 'F'])

    사용 예시:
        writer = CumulativeExcelWriter("template.xlsx", "all_results.xlsx")
        writer.append_result("exp_name", metrics, {"fusion_type": "mhca"})
    """

    HEADER_ROWS = 3  # 헤더 행 수 (1-indexed에서 데이터는 row 4부터)

    def __init__(self, template_path: str, output_path: str, classes=None):
        self.template_path = template_path
        self.output_path = output_path
        self.classes = classes or ['N', 'S', 'V', 'F']

        # 파일이 없으면 템플릿에서 생성
        if not os.path.exists(output_path):
            shutil.copy(template_path, output_path)
            print(f"[CumulativeExcel] Created new file from template: {output_path}")
        else:
            print(f"[CumulativeExcel] Using existing file: {output_path}")

        # 현재 마지막 row 찾기
        self.current_row = self._find_last_row()
        print(f"[CumulativeExcel] Next row: {self.current_row + 1}")

    def _find_last_row(self) -> int:
        """
        Performance Metrics 시트에서 마지막으로 기록된 row 찾기

        Returns:
            마지막 데이터 row (0-indexed)
        """
        wb = load_workbook(self.output_path)
        ws = wb['Performance Metrics']

        # A열(실험명)에서 마지막으로 값이 있는 row 찾기
        last_row = self.HEADER_ROWS  # 최소값 (헤더 다음)

        for row in range(self.HEADER_ROWS + 1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value is not None:
                last_row = row

        wb.close()
        return last_row  # 0-indexed로 변환하지 않음 (openpyxl은 1-indexed)

    def append_result(
        self,
        exp_name: str,
        metrics: dict,
        exp_config: dict = None,
        best_type: str = "auprc"
    ):
        """
        실험 결과를 다음 row에 추가

        Args:
            exp_name: 실험 이름
            metrics: calculate_metrics 결과 dict
            exp_config: 실험 설정 dict (fusion_type 등)
            best_type: 'auroc', 'auprc', 'recall' 등
        """
        wb = load_workbook(self.output_path)
        ws = wb['Performance Metrics']

        self.current_row += 1
        row = self.current_row

        # 실험명 (best_type + config 정보 포함)
        config_str = ""
        if exp_config:
            config_parts = []
            if exp_config.get('fusion_type'):
                config_parts.append(f"fusion={exp_config['fusion_type']}")
            if exp_config.get('fusion_num_heads'):
                config_parts.append(f"h={exp_config['fusion_num_heads']}")
            if config_parts:
                config_str = f"_{'_'.join(config_parts)}"

        timestamp = datetime.now().strftime("%m%d_%H%M")
        full_name = f"{exp_name}_{best_type}{config_str}_{timestamp}"
        ws.cell(row=row, column=1, value=full_name)

        # Macro metrics (columns 2-6)
        ws.cell(row=row, column=2, value=metrics.get('macro_accuracy', metrics.get('acc', 0)))
        ws.cell(row=row, column=3, value=metrics.get('macro_recall', 0))
        ws.cell(row=row, column=4, value=metrics.get('macro_specificity', 0))
        ws.cell(row=row, column=5, value=metrics.get('macro_precision', metrics.get('macro_prec', 0)))
        ws.cell(row=row, column=6, value=metrics.get('macro_f1', 0))

        # Weighted metrics (columns 7-11)
        ws.cell(row=row, column=7, value=metrics.get('weighted_accuracy', metrics.get('acc', 0)))
        ws.cell(row=row, column=8, value=metrics.get('weighted_recall', 0))
        ws.cell(row=row, column=9, value=metrics.get('weighted_specificity', 0))
        ws.cell(row=row, column=10, value=metrics.get('weighted_precision', metrics.get('weighted_prec', 0)))
        ws.cell(row=row, column=11, value=metrics.get('weighted_f1', 0))

        # Per-class metrics (N, S, V, F)
        per_class_acc = metrics.get('per_class_acc', metrics.get('per_class_accuracy', [0]*4))
        per_class_recall = metrics.get('per_class_recall', [0]*4)
        per_class_spec = metrics.get('per_class_specificity', [0]*4)
        per_class_prec = metrics.get('per_class_precision', [0]*4)
        per_class_f1 = metrics.get('per_class_f1', [0]*4)

        for i, cls in enumerate(self.classes):
            base_col = 12 + i * 5
            ws.cell(row=row, column=base_col, value=per_class_acc[i] if i < len(per_class_acc) else 0)
            ws.cell(row=row, column=base_col + 1, value=per_class_recall[i] if i < len(per_class_recall) else 0)
            ws.cell(row=row, column=base_col + 2, value=per_class_spec[i] if i < len(per_class_spec) else 0)
            ws.cell(row=row, column=base_col + 3, value=per_class_prec[i] if i < len(per_class_prec) else 0)
            ws.cell(row=row, column=base_col + 4, value=per_class_f1[i] if i < len(per_class_f1) else 0)

        wb.save(self.output_path)
        print(f"[CumulativeExcel] Result appended → row {row}: {full_name}")

    def append_confusion_matrix(self, exp_name: str, cm: np.ndarray, best_type: str = "auprc"):
        """
        Confusion 시트에 혼동 행렬 추가

        Args:
            exp_name: 실험 이름
            cm: 혼동 행렬 (4x4)
            best_type: 'auroc', 'auprc', 'recall' 등
        """
        wb = load_workbook(self.output_path)
        ws = wb['Confusion']

        # 마지막 confusion matrix 위치 찾기
        last_block_row = 1
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value is not None:
                last_block_row = row

        # 다음 블록 시작 위치 (8행 간격)
        block_start = ((last_block_row - 1) // 8 + 1) * 8 + 1

        full_name = f"{exp_name}_{best_type}"

        # EXPNAME 행
        ws.cell(row=block_start, column=1, value=full_name)

        # Predicted 헤더
        ws.cell(row=block_start + 1, column=3, value="Predicted")

        # 클래스 헤더
        for i, cls in enumerate(self.classes):
            ws.cell(row=block_start + 2, column=3 + i, value=cls)

        # Actual 라벨 및 데이터
        ws.cell(row=block_start + 3, column=1, value="Actual")
        for i, cls in enumerate(self.classes):
            ws.cell(row=block_start + 3 + i, column=2, value=cls)
            for j in range(len(self.classes)):
                ws.cell(row=block_start + 3 + i, column=3 + j, value=int(cm[i, j]))

        wb.save(self.output_path)
        print(f"[CumulativeExcel] Confusion matrix appended: {full_name}")

    def get_current_row(self) -> int:
        """현재 기록 위치 반환"""
        return self.current_row

    def get_record_count(self) -> int:
        """기록된 실험 수 반환"""
        return max(0, self.current_row - self.HEADER_ROWS)
